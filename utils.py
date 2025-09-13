import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import os
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import numpy as np

import pandas as pd
import geopandas as gpd
import pandas as pd
import matplotlib.pyplot as plt
from pathlib import Path

def ajustar_columnas_excel(ws):
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    if str(cell.value).startswith("=HYPERLINK"):
                        cell_length = 12
                    else:
                        cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def aplicar_validacion_checkbox(ws, column_name, num_rows):
    col_index = None
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == column_name:
            col_index = col
            break
    if col_index is None:
        return
    col_letter = get_column_letter(col_index)
    dv = DataValidation(
        type="list",
        formula1='"✅ Confirmado,❌ Falso Positivo,❌ Falso Negativo"',
        allow_blank=True,
    )
    rango = f"{col_letter}2:{col_letter}{num_rows + 1}"
    dv.add(rango)
    ws.add_data_validation(dv)


def normalizar_mundissec(codigo):
    return str(codigo).strip().zfill(11)


class MetricsAnalyzer:
    """Clase para analizar métricas de cycleway de forma unificada"""

    def __init__(self, config: Dict):
        """Inicializar analizador con configuración"""
        self.config = config
        self.setup_paths()

        # Definir métricas disponibles

        self.metricas_automaticas = ["Completeness (%)", "Accuracy (%)", "SCI"]
        self.metricas_validadas = ["Precision (%)", "Recall (%)", "F1_Score"]

    def setup_paths(self):
        """Configurar rutas de archivos y directorios"""
        # Crear solo la carpeta principal de salida

        Path("TABLAS_RESULTADOS_INTERPRETACIONES").mkdir(exist_ok=True)

    def load_territorial_data(self) -> pd.DataFrame:
        """Cargar datos territoriales (secciones, distritos, zonas)"""
        print("📍 Cargando datos territoriales...")

        try:
            # Cargar archivo principal de secciones

            df_secciones = pd.read_excel(
                self.config["archivo_secciones"], engine="openpyxl"
            )
            df_secciones["MUNDISSEC"] = df_secciones["MUNDISSEC"].astype(str)
            print(f"✅ Secciones base cargadas: {len(df_secciones)}")

            return df_secciones
        except Exception as e:
            print(f"❌ Error cargando datos territoriales: {e}")
            return pd.DataFrame()

    def load_automatic_metrics(self) -> pd.DataFrame:
        """Cargar y consolidar métricas automáticas de múltiples archivos"""
        print("📊 Cargando métricas automáticas...")

        carpeta_excel = Path(self.config["carpeta_resultados"])
        archivos_xlsx = list(carpeta_excel.glob("*.xlsx"))
        df_consolidado = pd.DataFrame()

        for archivo in archivos_xlsx:
            try:
                # Saltar archivos específicos

                if archivo.name in [
                    "metricas_validadas_post_GSV.xlsx",
                    "tabla_resumen_validacion.xlsx",
                ]:
                    continue
                df = pd.read_excel(archivo, engine="openpyxl")

                # Verificar que tenga las columnas necesarias

                if all(
                    col in df.columns for col in self.metricas_automaticas + ["Sección"]
                ):
                    df["Sección"] = df["Sección"].astype(str)
                    df_consolidado = pd.concat(
                        [df_consolidado, df], ignore_index=True)
                    print(f"✅ Procesado: {archivo.name}")
                else:
                    print(f"⏭️ Omitido: {archivo.name} (faltan columnas)")
            except Exception as e:
                print(f"❌ Error en {archivo.name}: {e}")
        if not df_consolidado.empty:
            # Convertir a numérico

            for col in self.metricas_automaticas:
                df_consolidado[col] = pd.to_numeric(
                    df_consolidado[col], errors="coerce"
                )
        print(
            f"📈 Métricas automáticas consolidadas: {len(df_consolidado)} registros")
        return df_consolidado

    def load_validated_metrics(self) -> Tuple[pd.DataFrame, pd.DataFrame]:
        """Cargar métricas validadas de archivos específicos"""
        print("🔍 Cargando métricas validadas...")

        df_validadas = pd.DataFrame()
        df_resumen_validacion = pd.DataFrame()

        # Cargar archivo de métricas validadas

        archivo_validadas = (
            Path(self.config["carpeta_resultados"]) /
            "metricas_validadas_post_GSV.xlsx"
        )
        if archivo_validadas.exists():
            try:
                df_validadas = pd.read_excel(archivo_validadas)
                df_validadas["Seccion"] = df_validadas["Seccion"].astype(str)

                # Convertir a numérico

                for col in self.metricas_validadas:
                    if col in df_validadas.columns:
                        df_validadas[col] = pd.to_numeric(
                            df_validadas[col], errors="coerce"
                        )
                print(
                    f"✅ Métricas validadas cargadas: {len(df_validadas)} registros")
            except Exception as e:
                print(f"❌ Error cargando métricas validadas: {e}")
        # Cargar tabla resumen de validación

        archivo_resumen = (
            Path(self.config["carpeta_resultados"]) /
            "tabla_resumen_validacion.xlsx"
        )
        if archivo_resumen.exists():
            try:
                df_resumen_validacion = pd.read_excel(archivo_resumen)
                df_resumen_validacion["Sección"] = df_resumen_validacion[
                    "Sección censal"
                ].astype(str)

                # Convertir campos numéricos

                cols_numericas = [
                    "# Commission confirmadas",
                    "# Commission",
                    "Precision (%)",
                    "Recall (%)",
                    "F1_Score",
                ]
                for col in cols_numericas:
                    if col in df_resumen_validacion.columns:
                        df_resumen_validacion[col] = pd.to_numeric(
                            df_resumen_validacion[col], errors="coerce"
                        )
                # Calcular % confirmadas

                df_resumen_validacion["% commissions confirmadas"] = (
                    df_resumen_validacion["# Commission confirmadas"]
                    / df_resumen_validacion["# Commission"]
                ) * 100

                print(
                    f"✅ Resumen validación cargado: {len(df_resumen_validacion)} registros"
                )
            except Exception as e:
                print(f"❌ Error cargando resumen validación: {e}")
        return df_validadas, df_resumen_validacion
    def analyze_by_district(
        self,
        df_metricas: pd.DataFrame,
        df_territorial: pd.DataFrame,
        tipo: str = "automaticas",
    ) -> pd.DataFrame:
        """Analizar métricas por distrito"""
        print(f"🏙️ Analizando por distrito ({tipo})...")

        if df_metricas.empty or df_territorial.empty:
            return pd.DataFrame()
        # Unir datos

        df_merged = df_metricas.merge(
            df_territorial, left_on="Sección", right_on="MUNDISSEC", how="left"
        )

        # Seleccionar métricas según tipo

        if tipo == "automaticas":
            cols_metricas = self.metricas_automaticas
        else:
            cols_metricas = self.metricas_validadas + \
                ["% commissions confirmadas"]
        # Agrupar por distrito

        agg_dict = {
            col: "mean" for col in cols_metricas if col in df_merged.columns}
        agg_dict["Sección"] = "count"

        df_distrito = (
            df_merged.groupby("NOM")
            .agg(agg_dict)
            .rename(columns={"Sección": "Nº de secciones"})
            .reset_index()
        )

        return df_distrito

    def analyze_by_zone(self, df_metricas: pd.DataFrame, df_territorial: pd.DataFrame) -> Dict[str, pd.DataFrame]:
        """
        Analiza métricas por subzonas:
        - Por distancia al centro urbano (cuartiles de proximidad).
        - Por densidad urbana (Alta/Baja densidad y cuartiles).
        Genera tres outputs:
        - resumen_metricas_por_zona_distancia.xlsx
        - resumen_metricas_por_zona_densidad.xlsx
        - asignacion_zonas_subzonas.xlsx
        """
        print("🌍 Analizando por distancia y por densidad...")

        zonas = self.config.get('archivos_zonas', {})
        resultados_distancia = []
        resultados_densidad = []
        resultados_asignacion = []

        columnas_densidad = ['DENS_POB', 'DENS_POB_KM2',
                             'DENSIDAD', 'DENSIDAD_POBLACIONAL']

        for archivo, zona_base in zonas.items():
            if Path(archivo).exists():
                print(f"   Cargando desde {archivo}")
                df_zona = pd.read_excel(archivo)
                df_zona.columns = [col.strip().upper()
                                   for col in df_zona.columns]
                df_zona["Sección"] = df_zona["MUNDISSEC"].astype(str)

                # === CLASIFICACIÓN POR DISTANCIA AL CENTRO ===
                if 'NEAR_DIST' in df_zona.columns and zona_base == "Distancia":
                    col_clasif = 'NEAR_DIST'
                    cuartil_labels = [
                        "25% más centrales",
                        "Intermedio-central",
                        "Intermedio-periférico",
                        "25% más periféricos"
                    ]
                    df_merge = df_zona.merge(
                        df_metricas, on="Sección", how="left")

                    if df_merge[col_clasif].notnull().sum() > 4:
                        try:
                            df_merge['Subzona'] = pd.qcut(
                                df_merge[col_clasif], 4, labels=cuartil_labels)
                        except Exception as e:
                            print(
                                f"   ⚠️ No se pudo segmentar en cuartiles (distancia): {e}")
                            df_merge['Subzona'] = "Sin clasificar"
                    else:
                        df_merge['Subzona'] = "Sin clasificar"

                    resumen = df_merge.groupby(['Subzona']).agg({
                        "Completeness (%)": "mean",
                        "Accuracy (%)": "mean",
                        "SCI": "mean",
                        "Sección": "count"
                    }).rename(columns={"Sección": "Nº secciones"}).reset_index()

                    for col in ["Completeness (%)", "Accuracy (%)", "SCI"]:
                        if col in resumen.columns:
                            resumen[col] = resumen[col].round(2)

                    resultados_distancia.append(resumen)
                    resultados_asignacion.append(
                        df_merge[['MUNDISSEC', 'Subzona']])
                    print(
                        f"   → Subzonas por distancia generadas: {len(resumen)}")

                # === CLASIFICACIÓN POR DENSIDAD ===
                elif zona_base == "Densidad":  # ← Cambiado para que coincida con  config
                    col_dens = next(
                        (c.upper() for c in columnas_densidad if c.upper() in df_zona.columns), None)

                    if col_dens:
                        cuartil_labels = [
                            "25% más densas",
                            "Densidad intermedia-alta",
                            "Densidad intermedia-baja",
                            "25% menos densas"
                        ]
                        df_zona["Zona"] = "Densidad"  # Asignamos zona base
                        df_merge = df_zona.merge(
                            df_metricas, on="Sección", how="left")

                        if df_merge[col_dens].notnull().sum() > 4:
                            try:
                                df_merge['Subzona'] = pd.qcut(
                                    df_merge[col_dens], 4, labels=cuartil_labels)
                            except Exception as e:
                                print(
                                    f"   ⚠️ No se pudo segmentar en cuartiles (densidad): {e}")
                                df_merge['Subzona'] = "Sin clasificar"
                        else:
                            df_merge['Subzona'] = "Sin clasificar"

                        resumen = df_merge.groupby(['Subzona']).agg({
                            "Completeness (%)": "mean",
                            "Accuracy (%)": "mean",
                            "SCI": "mean",
                            "Sección": "count"
                        }).rename(columns={"Sección": "Nº secciones"}).reset_index()

                        for col in ["Completeness (%)", "Accuracy (%)", "SCI"]:
                            if col in resumen.columns:
                                resumen[col] = resumen[col].round(2)

                        resultados_densidad.append(resumen)
                        resultados_asignacion.append(
                            df_merge[['MUNDISSEC', 'Zona', 'Subzona']])
                        print(
                            f"   → Subzonas por densidad agregadas: {len(resumen)}")
                    else:
                        print(
                            f"   ⚠️ Archivo {archivo} no contiene una columna de densidad reconocida")
                        print(
                            f"   Columnas disponibles: {list(df_zona.columns)}")

                else:
                    print(
                        f"   ⚠️ Tipo de zona no reconocido para {archivo}: {zona_base}")

            else:
                print(f"⚠️ Archivo no encontrado: {archivo}")

        outputs = {}
        df_zonas = pd.concat(
            resultados_distancia, ignore_index=True) if resultados_distancia else pd.DataFrame()
        df_zonas = df_zonas.drop_duplicates(subset="Subzona", keep="first")
        df_densidad = pd.concat(
            resultados_densidad, ignore_index=True) if resultados_densidad else pd.DataFrame()
        df_asignacion = pd.concat(
            resultados_asignacion, ignore_index=True) if resultados_asignacion else pd.DataFrame()
        if not df_asignacion.empty:
            df_asignacion["MUNDISSEC"] = df_asignacion["MUNDISSEC"].astype(
                str).str.zfill(11)

        outputs["resumen_metricas_por_zona_distancia"] = df_zonas
        outputs["resumen_metricas_por_zona_densidad"] = df_densidad
        outputs["asignacion_zonas_subzonas"] = df_asignacion

        print("\nResumen por subzonas (distancia al centro):")
        print(df_zonas)
        print("\nResumen por subzonas (densidad):")
        print(df_densidad)
        print("\nAsignación de MUNDISSEC a subzonas:")
        print(df_asignacion)

        return outputs

    def export_results(self, resultados: Dict[str, pd.DataFrame]):
        """Exportar todos los resultados a archivos Excel"""
        print("💾 Exportando resultados...")

        archivos_generados = []

        for nombre, df in resultados.items():
            if isinstance(df, pd.DataFrame) and not df.empty:
                # Todos los archivos van a la misma carpeta

                ruta = Path("TABLAS_RESULTADOS_INTERPRETACIONES") / \
                    f"{nombre}.xlsx"

                try:
                    df.to_excel(
                        ruta, index=False if "resumen" in nombre else True)
                    archivos_generados.append(str(ruta))
                    print(f"✅ Exportado: {ruta}")
                except Exception as e:
                    print(f"❌ Error exportando {nombre}: {e}")
        return archivos_generados
    def generar_tabla_resumen_distrital(
        path_metricas_con: str,
        path_metricas_sin: str,
        path_secciones_con: str,
        path_secciones_sin: str,
        output_excel: str = "TABLAS_RESULTADOS_INTERPRETACIONES/tabla_resumen_validacion_distrital.xlsx"
    ):
        """
        Genera una tabla resumen de métricas por distrito a partir de archivos ya calculados,
        usando métricas automáticas para secciones CON infraestructura y Precision visual para SIN.
        """
        import pandas as pd
        from pathlib import Path

        # Cargar archivos
        df_con = pd.read_excel(path_metricas_con)
        df_sin = pd.read_excel(path_metricas_sin)
        df_con_secc = pd.read_excel(path_secciones_con)
        df_sin_secc = pd.read_excel(path_secciones_sin)

        # Normalizar claves
        df_con["Seccion"] = df_con["Seccion"].astype(str).str.zfill(11)
        df_sin["Seccion"] = df_sin["Seccion"].astype(str).str.zfill(11)
        df_con_secc["MUNDISSEC"] = df_con_secc["MUNDISSEC"].astype(str).str.zfill(11)
        df_sin_secc["MUNDISSEC"] = df_sin_secc["MUNDISSEC"].astype(str).str.zfill(11)

        # Clasificación
        df_con["Tipo"] = "CON"
        df_sin["Tipo"] = "SIN"

        # Unir y simular columna de distrito con primeros 5 dígitos del código sección
        df_total = pd.concat([df_con, df_sin], ignore_index=True)
        df_total["Distrito"] = df_total["Seccion"].str[:5]

        # Agrupación por distrito
        resultados = []

        for distrito, grupo in df_total.groupby("Distrito"):
            grupo_con = grupo[grupo["Tipo"] == "CON"]
            grupo_sin = grupo[grupo["Tipo"] == "SIN"]

            n_con = len(grupo_con)
            n_sin = len(grupo_sin)

            accuracy = grupo_con["Accuracy (%)"].mean() if not grupo_con.empty else None
            completeness = grupo_con["Completeness (%)"].mean() if not grupo_con.empty else None
            sci = grupo_con["SCI (%)"].mean() if "SCI (%)" in grupo_con.columns and not grupo_con.empty else None
            precision = grupo_sin["Precision (%)"].mean() if not grupo_sin.empty else None

            resultados.append({
                "Distrito": distrito,
                "Accuracy CON (%)": round(accuracy, 2) if accuracy else None,
                "Completeness CON (%)": round(completeness, 2) if completeness else None,
                "SCI CON (%)": round(sci, 2) if sci else None,
                "Precision SIN (%)": round(precision, 2) if precision else None,
                "Secciones CON": n_con,
                "Secciones SIN": n_sin,
                "Total Secciones": n_con + n_sin
            })

        df_resultado = pd.DataFrame(resultados).sort_values("Distrito")

        # Exportar Excel
        Path(output_excel).parent.mkdir(parents=True, exist_ok=True)
        df_resultado.to_excel(output_excel, index=False)
        print(f"✅ Tabla exportada a: {output_excel}")

        return df_resultado

    def run_complete_analysis(self):
        """Ejecutar análisis completo"""
        print("🚀 Iniciando análisis completo de métricas...")
        print("=" * 60)

        # 1. Cargar datos base

        df_territorial = self.load_territorial_data()
        df_auto = self.load_automatic_metrics()
        df_val, df_resumen_val = self.load_validated_metrics()

        # 2. Generar análisis

        resultados = {}

        # Estadísticas generales

        if not df_auto.empty:
            zonas_outputs = self.analyze_by_zone(df_auto, df_territorial)
            resultados.update(zonas_outputs)
        # Análisis por distrito

        if not df_auto.empty and not df_territorial.empty:
            resultados["metricas_por_distrito_prevalidacion"] = (
                self.analyze_by_district(
                    df_auto, df_territorial, "automaticas")
            )
        if not df_resumen_val.empty and not df_territorial.empty:
            resultados["metricas_por_distrito_postvalidacion"] = (
                self.analyze_by_district(
                    df_resumen_val, df_territorial, "validadas")
            )
        # Análisis por zona

        if not df_auto.empty:
            resultados["resumen_metricas_por_zona"] = self.analyze_by_zone(
                df_auto, df_territorial
            )
        # 3. Exportar resultados

        archivos_generados = self.export_results(resultados)

        # 4. Mostrar resumen final

        print("\n" + "=" * 60)
        print("✅ ANÁLISIS COMPLETADO")
        print("=" * 60)

        if archivos_generados:
            print(
                f"📁 Archivos generados en 'TABLAS_RESULTADOS_INTERPRETACIONES' ({len(archivos_generados)}):"
            )
            for archivo in archivos_generados:
                print(f"   • {Path(archivo).name}")
        else:
            print("⚠️ No se generaron archivos debido a errores en el procesamiento.")
        print("\n🎯 Análisis ejecutado exitosamente.")

        return resultados


def generar_tabla_resumen_distrital(
    path_postvalidacion: str = "./resultados_cycleway/metricas_validadas_post_GSV.xlsx",
    path_prevalidacion: str = "./resultados_cycleway/metricas_automaticas_prevalidacion.xlsx",
    path_secciones_con: str = "./resultados_cycleway/secciones_con_oficial.xlsx",
    path_secciones_sin: str = "./resultados_cycleway/secciones_sin_oficial.xlsx",
    path_territorial: str = "./SECCIONS_CENSALS_DIFFERENC/TODAS_SECCIONES.xlsx",
    output_excel: str = "TABLAS_RESULTADOS_INTERPRETACIONES/tabla_resumen_validacion_distrital.xlsx"
):
    """
    Genera una tabla resumen de métricas post-validación por distrito.
    Incluye métricas ajustadas con Partial Match para Accuracy y Completeness,
    y calcula la Precision (%) para secciones SIN infraestructura oficial.
    También señala si el distrito fue validado visualmente.
    """

    df_post = pd.read_excel(path_postvalidacion)
    df_pre = pd.read_excel(path_prevalidacion)
    df_con = pd.read_excel(path_secciones_con)
    df_sin = pd.read_excel(path_secciones_sin)
    df_territorial = pd.read_excel(path_territorial)

    # Normalización
    if "Seccion" not in df_post.columns:
        if "Sección" in df_post.columns:
            df_post.rename(columns={"Sección": "Seccion"}, inplace=True)
        else:
            print("❗ Columnas disponibles en df_post:", df_post.columns.tolist())
            raise KeyError("❌ La columna 'Seccion' no existe en el archivo de postvalidación.")

    if "Seccion" not in df_pre.columns:
        if "Sección" in df_pre.columns:
            df_pre.rename(columns={"Sección": "Seccion"}, inplace=True)
        else:
            print("❗ Columnas disponibles en df_pre:", df_pre.columns.tolist())
            raise KeyError("❌ La columna 'Seccion' no existe en el archivo de prevalidación.")

    df_post["Seccion"] = df_post["Seccion"].astype(str).str.zfill(11)
    df_pre["Seccion"] = df_pre["Seccion"].astype(str).str.zfill(11)
    df_con["MUNDISSEC"] = df_con["MUNDISSEC"].astype(str).str.zfill(11)
    df_sin["MUNDISSEC"] = df_sin["MUNDISSEC"].astype(str).str.zfill(11)
    df_territorial["MUNDISSEC"] = df_territorial["MUNDISSEC"].astype(str).str.zfill(11)

    df_post = df_post.merge(df_territorial, left_on="Seccion", right_on="MUNDISSEC", how="left")
    df_pre = df_pre.merge(df_territorial, left_on="Seccion", right_on="MUNDISSEC", how="left")

    secciones_con = set(df_con["MUNDISSEC"])
    secciones_sin = set(df_sin["MUNDISSEC"])

    def clasificar(seccion):
        if seccion in secciones_con:
            return "CON"
        elif seccion in secciones_sin:
            return "SIN"
        return "DESCONOCIDO"

    df_post["Tipo"] = df_post["Seccion"].apply(clasificar)
    df_pre["Tipo"] = df_pre["Seccion"].apply(clasificar)

    resultados = []
    distritos = df_post["NOM"].dropna().unique()

    for distrito in sorted(distritos):
        post_d = df_post[df_post["NOM"] == distrito]
        pre_d = df_pre[df_pre["NOM"] == distrito]

        con_post = post_d[post_d["Tipo"] == "CON"]
        con_pre = pre_d[pre_d["Tipo"] == "CON"]
        sin_post = post_d[post_d["Tipo"] == "SIN"]

        tp = con_post["True_Positives"].sum()
        fp = con_post["False_Positives"].sum()
        fn = con_post["False_Negatives"].sum()
        pm = con_pre["Partial Match"].sum() if "Partial Match" in con_pre.columns else 0
        sci = con_pre["SCI"].mean()

        acc = (tp + pm) / (tp + pm + fp + fn) * 100 if (tp + pm + fp + fn) else None
        comp = (tp + pm) / (tp + pm + fn) * 100 if (tp + pm + fn) else None

        tp_sin = sin_post["True_Positives"].sum()
        fp_sin = sin_post["False_Positives"].sum()
        precision = tp_sin / (tp_sin + fp_sin) * 100 if (tp_sin + fp_sin) else None

        estado = "Validado" if not sin_post.empty else "Solo CON"

        resultados.append({
            "Distrito": distrito,
            "Accuracy CON (%)": round(acc, 2) if acc is not None else None,
            "Completeness CON (%)": round(comp, 2) if comp is not None else None,
            "SCI CON (%)": round(sci, 2) if sci is not None else None,
            "Precision SIN (%)": round(precision, 2) if precision is not None else None,
            "Secciones CON": len(con_post),
            "Secciones SIN": len(sin_post),
            "Total": len(post_d),
            "Estado Validación": estado
        })

    df_resultado = pd.DataFrame(resultados)
    Path(output_excel).parent.mkdir(parents=True, exist_ok=True)
    df_resultado.to_excel(output_excel, index=False)
    print(f"✅ Tabla exportada a: {output_excel}")

    return df_resultado


def main():
    """Función principal"""
    # Configuración

    config = {
        "carpeta_resultados": "resultados_cycleway",
        "archivo_secciones": "SECCIONS_CENSALS_DIFFERENC/TODAS_SECCIONES.xlsx",
        "archivos_zonas": {
            "SECCIONS_CENSALS_DIFFERENC/SECCIONES_CENTRO.xlsx": "Distancia",
            "SECCIONS_CENSALS_DIFFERENC/SECCIONES_DENSIDAD.xlsx": "Densidad",
        },
    }

    # Ejecutar análisis

    analyzer = MetricsAnalyzer(config)
    resultados = analyzer.run_complete_analysis()
    return resultados

# generar_tabla_resumen_distrital()

if __name__ == "__main__":
    main()
