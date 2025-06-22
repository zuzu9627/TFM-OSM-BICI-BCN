import hashlib
import logging
import os
import re
import sys
import warnings
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple

# Third-party libraries
import geopandas as gpd
import matplotlib.pyplot as plt
import numpy as np
import openpyxl 
import pandas as pd
from ohsome import OhsomeClient
from shapely.errors import ShapelyDeprecationWarning
from shapely.geometry import Point

# openpyxl helpers
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
# Custom
from utils import ajustar_columnas_excel, aplicar_validacion_checkbox, normalizar_mundissec
# Configurar warnings y logging

warnings.filterwarnings("ignore", category=ShapelyDeprecationWarning)
logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)


@dataclass
class Config:
    """Configuración centralizada del análisis"""
    # Directorios y archivos

    script_dir: Path = Path(__file__).parent
    output_dir: Path = script_dir / "resultados_cycleway"

    # Archivos de entrada

    archivos_xlsx: List[str] = None
    path_secciones: str = (
        "./SECCIONS_CENSALS_DIFFERENC/DISTR+SECCIONS+POB_2024/EXPORTED_2024_BICI_BCN_Layer.shp"
        # "./UNITATS_CENSALS_2024/unitats_censals_2024.shp"
    )

    path_oficial: str = (
        "./DATOS_OFICIAL_CARRIL_BICI/2024_1T_CARRIL_BICI/2024_1T_CARRIL_BICI.shp"
    )
    # Parámetros de análisis

    años: List[str] = None
    buffer_tolerancia: int = 6
    usar_dummies: bool = False
    tag_filter: str = "type:way and (highway=cycleway)"

    # Umbrales de calidad

    umbral_completeness: float = 85.0
    umbral_sci: float = 80.0
    umbral_accuracy: float = 75.0

    # Paralelización

    max_workers: int = 4

    def __post_init__(self):
        if self.archivos_xlsx is None:
            base_path = "./SECCIONS_CENSALS_DIFFERENC"
            self.archivos_xlsx = [
                f"{base_path}/TODAS_SECCIONES.xlsx"
            ]
        if self.años is None:
            self.años = ["2024"]
        self.output_dir.mkdir(exist_ok=True)


@dataclass
class ResultadosAnalisis:
    """Contenedor para resultados de análisis por sección"""

    seccion: str
    año: str
    partial_match: gpd.GeoDataFrame
    omission: gpd.GeoDataFrame
    commission: gpd.GeoDataFrame
    metricas: Dict[str, float]


class OptimizedCyclewayAnalyzer:
    """Analizador optimizado de ciclovías OSM vs datos oficiales con validación manual"""

    def __init__(self, config: Config):
        self.config = config
        self.client = OhsomeClient()
        self.shp_secciones = None
        self.shp_oficial = None
        self._cargar_shapefiles()

    def _cargar_shapefiles(self) -> None:
        """Carga y prepara los shapefiles base"""
        logger.info("Cargando shapefiles base...")
        try:
            self.shp_secciones = gpd.read_file(self.config.path_secciones).to_crs(
                "EPSG:4326"
            )
            self.shp_secciones["MUNDISSEC"] = (
                self.shp_secciones["MUNDISSEC"].astype(
                    str).str.strip().str.zfill(11)
            )
            self.shp_oficial = gpd.read_file(self.config.path_oficial).to_crs(
                "EPSG:4326"
            )
            logger.info(
                "Shapefiles cargados: %d secciones, %d tramos oficiales",
                len(self.shp_secciones),
                len(self.shp_oficial),
            )
        except Exception as e:
            logger.error("Error cargando shapefiles: %s", e)
            raise

    def obtener_secciones_unicas(self) -> Set[str]:
        """Extrae secciones censales únicas de todos los archivos Excel"""
        secciones = set()
        for archivo in self.config.archivos_xlsx:
            try:
                df = pd.read_excel(archivo)
                df["MUNDISSEC"] = df["MUNDISSEC"].apply(normalizar_mundissec)
                secciones.update(df["MUNDISSEC"].unique())
            except Exception as e:
                logger.warning("Error procesando %s: %s", archivo, e)
        logger.info("Total secciones censales únicas: %d", len(secciones))
        return sorted(secciones)

    def clasificar_secciones(self, secciones: List[str]) -> Tuple[List[str], List[str]]:
        """Clasifica secciones según presencia de geometría oficial y guarda resultados en Excel"""
        secciones_con_oficial, secciones_sin_oficial = [], []

        # Normalizar y filtrar secciones válidas
        secciones_norm = [normalizar_mundissec(s) for s in secciones]
        secciones_validas = self.shp_secciones[self.shp_secciones["MUNDISSEC"].isin(
            secciones_norm)]

        # Advertir sobre secciones no encontradas
        encontradas = set(secciones_validas["MUNDISSEC"])
        for s in secciones_norm:
            if s not in encontradas:
                logger.warning("Sección %s no encontrada en shapefile", s)

        # Clasificar cada sección válida
        for _, row in secciones_validas.iterrows():
            seccion = row["MUNDISSEC"]
            tiene_oficial = self.shp_oficial.intersects(row.geometry).any()
            if tiene_oficial:
                secciones_con_oficial.append(seccion)
            else:
                secciones_sin_oficial.append(seccion)

        # Log resultados
        logger.info("Secciones CON geometría oficial (%d): %s",
                    len(secciones_con_oficial), ", ".join(secciones_con_oficial))
        logger.info("Secciones SIN geometría oficial (%d): %s",
                    len(secciones_sin_oficial), ", ".join(secciones_sin_oficial))

        # === NUEVO: Guardar en Excel ===
        pd.DataFrame({'MUNDISSEC': secciones_con_oficial}).to_excel(
            str(self.config.output_dir / 'secciones_con_oficial.xlsx'), index=False)
        pd.DataFrame({'MUNDISSEC': secciones_sin_oficial}).to_excel(
            str(self.config.output_dir / 'secciones_sin_oficial.xlsx'), index=False)
        # ==============================
        return secciones_con_oficial, secciones_sin_oficial

    def procesar_seccion_sin_oficial(self, seccion: str) -> list:
        """
        Procesa una sección censal SIN geometría oficial.
        Exporta cualquier tramo OSM (commission) a archivo de validación visual solo si existen,
        y añade las secciones sin datos a self.secciones_sin_resultados.
        """
        logger.info(f"Procesando sección SIN oficial {seccion}...")

        seccion = normalizar_mundissec(seccion)
        zona = self.shp_secciones[self.shp_secciones["MUNDISSEC"] == seccion]
        if zona.empty:
            logger.warning(f"Sección {seccion} no encontrada en shapefile")
            return []

        bbox = ",".join(map(str, zona.total_bounds))
        resultados_seccion = []

        for año in self.config.años:
            logger.info(f"Procesando año {año} para sección {seccion}")

            # Descargar y recortar OSM
            gdf_osm = self.descargar_osm_ohsome(bbox, f"{año}-07-01")
            if not gdf_osm.empty:
                gdf_osm = gpd.overlay(gdf_osm.reset_index(drop=True),
                                      zona.reset_index(drop=True)[
                    ["geometry"]],
                    how="intersection", keep_geom_type=False)
                gdf_osm = self._limpiar_columnas_geometria(gdf_osm)

            # Procesar resultados
            if not gdf_osm.empty:
                # Crear directorio y guardar archivos
                output_dir = self.config.output_dir / f"seccion_{seccion}"
                output_dir.mkdir(exist_ok=True)

                self.generar_validacion_unificada(gpd.GeoDataFrame(), gpd.GeoDataFrame(),
                                                  gdf_osm, output_dir, año, seccion)

                metricas = {"Partial Match (Buffer) (n)": 0, "Omission (n)": 0,
                            "Commission (n)": len(gdf_osm), "Accuracy (%)": 0.0,
                            "Completeness (%)": 0.0, "SCI": 0.0,
                            "Longitud oficial (m)": 0.0, "Longitud representada (m)": 0.0}

                pd.DataFrame([metricas]).to_excel(
                    output_dir / f"metricas_{seccion}_{año}.xlsx", index=False)

                resultados = {"Partial Match (Buffer)": gpd.GeoDataFrame(),
                              "Omission": gpd.GeoDataFrame(), "Commission": gdf_osm}
                self._guardar_resultados_geoespaciales(
                    resultados, output_dir, año, seccion)

                resultados_seccion.append(ResultadosAnalisis(
                    seccion=seccion, año=año, partial_match=gpd.GeoDataFrame(),
                    omission=gpd.GeoDataFrame(), commission=gdf_osm, metricas=metricas))
            else:
                # Añadir a secciones sin resultados
                if not hasattr(self, "secciones_sin_resultados"):
                    self.secciones_sin_resultados = []
                if seccion not in self.secciones_sin_resultados:
                    self.secciones_sin_resultados.append(seccion)

        return resultados_seccion

    def descargar_osm_ohsome(self, bbox: str, fecha: str) -> gpd.GeoDataFrame:
        """Descarga datos OSM usando API Ohsome con manejo de errores mejorado"""
        try:
            response = self.client.elements.geometry.post(
                bboxes=bbox, time=fecha, filter=self.config.tag_filter, properties="tags,metadata")

            gdf = response._as_geodataframe()
            if gdf.empty:
                return gdf

            # Limpieza, validación y filtrado en una sola operación
            mask = (gdf.geometry.notnull() & gdf.is_valid &
                    gdf.geometry.geom_type.isin(["LineString", "MultiLineString"]))
            gdf = gdf[mask].set_crs("EPSG:4326")

            # Generar ID único basado en geometría
            gdf["@id"] = gdf.geometry.apply(
                lambda g: hashlib.md5(g.wkb).hexdigest())

            return gdf
        except Exception as e:
            logger.error(f"Error descargando datos OSM: {e}")
            return gpd.GeoDataFrame()

    def identificar_secciones_sin_oficial_con_commission(self, secciones_sin_oficial: List[str]) -> None:
        """Genera un resumen del número de Commission y Omission en las secciones sin oficial."""

        # Buscar archivos de validación
        archivos = (list(self.config.output_dir.glob("seccion_*/validacion_visual_unificada_*.xlsx")) +
                    list(self.config.output_dir.glob("seccion_*/dummy_validacion_visual_unificada_*.xlsx")))

        conteo_por_mundissec = {}

        for archivo in archivos:
            try:
                df = pd.read_excel(archivo)
                if df.empty:
                    continue

                col_categoria = next(
                    (c for c in df.columns if c.lower() == "categoria"), None)
                if col_categoria is None:
                    continue

                # Procesar por MUNDISSEC o extraer del nombre del archivo
                if "MUNDISSEC" in df.columns:
                    for cod in df["MUNDISSEC"].astype(str).str.zfill(11).unique():
                        df_seccion = df[df["MUNDISSEC"].astype(
                            str).str.zfill(11) == cod]
                        self._actualizar_conteo(
                            conteo_por_mundissec, cod, df_seccion, col_categoria)
                else:
                    match = re.search(r"seccion_(\d{11})", archivo.name)
                    cod = match.group(1) if match else "SIN_CODIGO"
                    self._actualizar_conteo(
                        conteo_por_mundissec, cod, df, col_categoria)

            except Exception as e:
                logger.warning(f"Error leyendo {archivo}: {e}")

        # Crear y guardar resumen
        df_resumen = pd.DataFrame([
            {"MUNDISSEC": cod, "Num Commission": v["Commission"],
             "Num Omission": v["Omission"], "Total": v["Commission"] + v["Omission"]}
            for cod, v in conteo_por_mundissec.items()
        ]).sort_values("MUNDISSEC")

        # Guardar archivos
        archivo_excel = self.config.output_dir / \
            "secciones_sin_oficial_con_commissions.xlsx"
        df_resumen.to_excel(archivo_excel, index=False)
        self._formatear_excel(archivo_excel)

        archivo_txt = self.config.output_dir / "secciones_sin_oficial_con_commission.txt"
        self._guardar_txt_resumen(archivo_txt, conteo_por_mundissec)

        # Mostrar resumen
        print("Resumen de commission y omission (sin oficial):")
        print(df_resumen.to_string(index=False))
        print(
            f"\nArchivos generados:\n- Excel: {archivo_excel}\n- TXT: {archivo_txt}")

    def _actualizar_conteo(self, conteo_por_mundissec: dict, cod: str, df: pd.DataFrame, col_categoria: str):
        """Actualiza el conteo de commission y omission para una sección."""
        if cod not in conteo_por_mundissec:
            conteo_por_mundissec[cod] = {"Commission": 0, "Omission": 0}

        conteo_por_mundissec[cod]["Commission"] += (
            df[col_categoria] == "Commission").sum()
        conteo_por_mundissec[cod]["Omission"] += (
            df[col_categoria] == "Omission").sum()

    def _formatear_excel(self, archivo_excel):
        """Aplica formato automático al archivo Excel."""

        wb = load_workbook(archivo_excel)
        ws = wb.active

        # Ajustar ancho de columnas y formatear
        for column in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in column)
            ws.column_dimensions[column[0].column_letter].width = min(
                max_length + 3, 30)

        # Formatear encabezados
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Centrar datos numéricos
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.alignment = Alignment(horizontal="center")

        wb.save(archivo_excel)
        wb.close()

    def _guardar_txt_resumen(self, archivo_txt, conteo_por_mundissec: dict):
        """Guarda el resumen en formato TXT."""
        total_comm = sum(v["Commission"]
                         for v in conteo_por_mundissec.values())
        total_omis = sum(v["Omission"] for v in conteo_por_mundissec.values())

        with open(archivo_txt, "w", encoding="utf-8") as f:
            f.write("RESUMEN DE COMMISSION Y OMISSION (SECCIONES SIN OFICIAL)\n")
            f.write("=" * 60 + "\n\n")
            f.write(f"TOTALES GENERALES:\nTotal Commission: {total_comm}\n")
            f.write(
                f"Total Omission: {total_omis}\nTotal General: {total_comm + total_omis}\n\n")
            f.write("DETALLE POR SECCIÓN:\n" + "-" * 40 + "\n")

            for cod in sorted(conteo_por_mundissec.keys()):
                v = conteo_por_mundissec[cod]
                total_seccion = v["Commission"] + v["Omission"]
                f.write(
                    f"{cod}: Commission={v['Commission']}, Omission={v['Omission']}, Total={total_seccion}\n")

    def comparar_osm_vs_oficial(self, gdf_osm: gpd.GeoDataFrame, ciclovias_oficial: gpd.GeoDataFrame) -> Dict[str, gpd.GeoDataFrame]:
        """Comparación espacial optimizada entre OSM y datos oficiales"""
        if gdf_osm.empty or ciclovias_oficial.empty:
            return {
                "Partial Match (Buffer)": gpd.GeoDataFrame(),
                "Omission": ciclovias_oficial.copy() if not ciclovias_oficial.empty else gpd.GeoDataFrame(),
                "Commission": gdf_osm.copy() if not gdf_osm.empty else gpd.GeoDataFrame(),
            }

        # Proyectar a sistema métrico
        ref_oficial = ciclovias_oficial.to_crs(epsg=25831).copy()
        comp_osm = gdf_osm.to_crs(epsg=25831).copy()

        # Preparar datos con vectorización
        ref_oficial["longitud"] = ref_oficial.geometry.length
        ref_oficial["longitud_buffer"] = 0.0
        ref_oficial["categoria"] = "Omission"
        ref_oficial["buffer"] = ref_oficial.geometry.buffer(
            self.config.buffer_tolerancia)
        comp_osm["usado"] = False

        # Análisis con índice espacial
        spatial_index = comp_osm.sindex
        for idx, row in ref_oficial.iterrows():
            candidates_idx = list(
                spatial_index.intersection(row["buffer"].bounds))
            if candidates_idx:
                candidates = comp_osm.iloc[candidates_idx]
                cercanos = candidates[candidates.geometry.intersects(
                    row["buffer"])]

                if not cercanos.empty:
                    ref_oficial.at[idx, "categoria"] = "Partial Match (Buffer)"
                    intersections = [row["buffer"].intersection(
                        geom) for geom in cercanos.geometry]
                    ref_oficial.at[idx, "longitud_buffer"] = sum(
                        i.length for i in intersections if not i.is_empty)
                    comp_osm.loc[cercanos.index, "usado"] = True

        # Generar resultados
        partial_match = ref_oficial[ref_oficial["categoria"]
                                    == "Partial Match (Buffer)"].copy()
        omission = ref_oficial[ref_oficial["categoria"] == "Omission"].copy()
        commission = comp_osm[~comp_osm["usado"]].copy()
        commission["categoria"] = "Commission"

        return {
            "Partial Match (Buffer)": partial_match,
            "Omission": omission,
            "Commission": commission
        }

    def calcular_metricas(self, resultados: Dict[str, gpd.GeoDataFrame], ciclovias_oficial: gpd.GeoDataFrame) -> Dict[str, float]:
        """Cálculo optimizado de métricas de calidad para secciones con infraestructura oficial evaluable"""

        # Conteos por categoría
        partial_match_count = len(resultados["Partial Match (Buffer)"])
        omission_count = len(resultados["Omission"])
        commission_count = len(resultados["Commission"])

        # ❌ Excluir secciones sin referencia oficial (solo comisiones)
        if partial_match_count == 0 and omission_count == 0:
            print(
                f"❌ Sección {codigo_seccion} excluida: solo contiene Commission (sin coincidencias ni omisiones)")
            return None

        # Longitudes en metros (EPSG:25831)
        longitud_buffer = resultados["Partial Match (Buffer)"].to_crs(
            "EPSG:25831").geometry.length.sum() if partial_match_count > 0 else 0
        longitud_oficial = ciclovias_oficial.to_crs(
            "EPSG:25831").geometry.length.sum() if not ciclovias_oficial.empty else 0

        # Métricas básicas
        reference_elements = partial_match_count + omission_count
        total_elements = reference_elements + commission_count

        completeness = (partial_match_count / reference_elements *
                        100) if reference_elements > 0 else 0
        accuracy = (partial_match_count / total_elements *
                    100) if total_elements > 0 else 0
        sci = (longitud_buffer / longitud_oficial) if longitud_oficial > 0 else 0

        return {
            "Partial Match (Buffer) (n)": partial_match_count,
            "Omission (n)": omission_count,
            "Commission (n)": commission_count,
            "Accuracy (%)": round(accuracy, 2),
            "Completeness (%)": round(completeness, 2),
            "SCI": round(sci, 3),
            "Longitud oficial (m)": round(longitud_oficial, 1),
            "Longitud representada (m)": round(longitud_buffer, 1),
        }

    def generar_validacion_unificada(self, partial_match: gpd.GeoDataFrame, omission: gpd.GeoDataFrame,
                                     commission: gpd.GeoDataFrame, output_dir: Path, año: str, seccion: str) -> None:
        """Genera un único archivo de validación visual por año y sección."""

        registros = []

        # Procesar Commission y Omission
        for categoria, gdf in [("Commission", commission), ("Omission", omission)]:
            for idx, row in gdf.iterrows():
                if row.geometry.is_empty:
                    continue

                # Convertir a EPSG:4326 si es necesario
                geom = (gpd.GeoSeries([row.geometry], crs=gdf.crs).to_crs("EPSG:4326").iloc[0]
                        if gdf.crs != "EPSG:4326" else row.geometry)

                # Obtener punto central
                punto = (geom.interpolate(0.5, normalized=True).centroid
                         if geom.geom_type.startswith("Multi") else geom.centroid)

                # Obtener OSM ID
                osm_id = self._obtener_osm_id(row, idx)
                if categoria == "Omission" and osm_id.isdigit():
                    osm_id = f"omission_{osm_id}"

                registros.append({
                    "ID": osm_id,
                    "Categoria": categoria,
                    "Lat": round(punto.y, 5),
                    "Lon": round(punto.x, 5),
                    "Link GSV": f'=HYPERLINK("https://www.google.com/maps/@?api=1&map_action=pano&viewpoint={punto.y},{punto.x}", "Ver GSV")',
                    "Imagen satélite": "Google Maps / ICGC",
                    "¿Es realmente un carril bici?": "",
                    "Observaciones": "",
                })

        if not registros:
            return

        # Crear DataFrame y guardar
        df = pd.DataFrame(registros)
        df["MUNDISSEC"] = seccion
        df["Año"] = año

        if self.config.usar_dummies:
            df = self._aplicar_etiquetas_dummy(df)
            nombre_archivo = f"dummy_validacion_visual_unificada_{año}_seccion_{seccion}.xlsx"
        else:
            nombre_archivo = f"validacion_visual_unificada_{año}_seccion_{seccion}.xlsx"

        # Guardar y aplicar formato
        path_salida = output_dir / nombre_archivo
        df.to_excel(path_salida, index=False)

        # Aplicar validación de datos

        wb = load_workbook(path_salida)
        ws = wb.active

        # Encontrar columna de validación
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == "¿Es realmente un carril bici?":
                aplicar_validacion_checkbox(
                    ws, "¿Es realmente un carril bici?", len(df))
                ajustar_columnas_excel(ws)
                break

        wb.save(path_salida)
        wb.close()

        logger.info(
            f"Archivo de validación unificada generado: {nombre_archivo} ({len(registros)} registros)")

    def _obtener_osm_id(self, row: pd.Series, idx_fallback: int) -> str:
        """Obtiene ID de OSM de la fila, con fallback al índice"""
        for id_col in ["@id", "id", "osm_id"]:
            if id_col in row.index and pd.notna(row[id_col]):
                return str(row[id_col])
        return str(idx_fallback)

    def procesar_seccion(self, seccion: str) -> list:
        """
        Procesa una sección censal comparando OSM y oficial, generando resultados sólo si hay información relevante.
        Añade las secciones sin datos a self.secciones_sin_resultados.
        """
        logger.info(f"Procesando sección censal {seccion}...")

        seccion = normalizar_mundissec(seccion)
        zona = self.shp_secciones[self.shp_secciones["MUNDISSEC"] == seccion]

        if zona.empty:
            logger.warning(f"Sección {seccion} no encontrada en shapefile")
            return []

        # Inicializar lista de secciones sin resultados si no existe
        if not hasattr(self, "secciones_sin_resultados"):
            self.secciones_sin_resultados = []

        bbox = ",".join(map(str, zona.total_bounds))
        resultados_seccion = []

        def _procesar_geodf(gdf, zona_geom):
            """Función auxiliar para procesar geodataframes con overlay"""
            if gdf.empty:
                return gdf
            return self._limpiar_columnas_geometria(
                gpd.overlay(gdf.reset_index(drop=True),
                            zona.reset_index(drop=True)[["geometry"]],
                            how="intersection", keep_geom_type=False)
            )

        for año in self.config.años:
            logger.info(f"Procesando año {año} para sección {seccion}")

            # Procesar datos OSM y oficiales
            gdf_osm = _procesar_geodf(
                self.descargar_osm_ohsome(bbox, f"{año}-07-01"), zona)
            ciclovias_oficial = _procesar_geodf(
                self.shp_oficial[self.shp_oficial.intersects(
                    zona.geometry.iloc[0])].copy(),
                zona
            )

            # Comparar y evaluar
            resultados = self.comparar_osm_vs_oficial(
                gdf_osm, ciclovias_oficial)

            if any(not df.empty for df in [resultados["Partial Match (Buffer)"],
                                           resultados["Omission"],
                                           resultados["Commission"]]):

                output_dir_seccion = self.config.output_dir / \
                    f"seccion_{seccion}"
                output_dir_seccion.mkdir(exist_ok=True)

                # Guardar todos los resultados
                self._guardar_resultados_geoespaciales(
                    resultados, output_dir_seccion, año, seccion)
                self.generar_validacion_unificada(
                    resultados["Partial Match (Buffer)"], resultados["Omission"],
                    resultados["Commission"], output_dir_seccion, año, seccion
                )

                # Calcular métricas y guardar Excel
                metricas = self.calcular_metricas(
                    resultados, ciclovias_oficial)
                excel_path = output_dir_seccion / \
                    f"metricas_{seccion}_{año}.xlsx"

                with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
                    pd.DataFrame([metricas]).to_excel(
                        writer, sheet_name="Metricas", index=False)
                    # Auto-ajustar ancho de columnas
                    for col in writer.sheets["Metricas"].columns:
                        max_len = max(len(str(cell.value))
                                      if cell.value else 0 for cell in col)
                        writer.sheets["Metricas"].column_dimensions[col[0].column_letter].width = min(
                            max_len + 2, 50)

                # Añadir resultado al resumen
                resultados_seccion.append(ResultadosAnalisis(
                    seccion=seccion, año=año,
                    partial_match=resultados["Partial Match (Buffer)"],
                    omission=resultados["Omission"],
                    commission=resultados["Commission"],
                    metricas=metricas
                ))
            else:
                # Añadir sección sin datos (evitar duplicados)
                if seccion not in self.secciones_sin_resultados:
                    self.secciones_sin_resultados.append(seccion)

        return resultados_seccion

    def _limpiar_columnas_geometria(self, gdf: gpd.GeoDataFrame) -> gpd.GeoDataFrame:
        """Elimina columnas de geometría duplicadas"""
        for col in gdf.columns:
            if col != "geometry" and gdf[col].dtype.name == "geometry":
                gdf = gdf.drop(columns=[col])
        gdf.set_geometry("geometry", inplace=True)
        return gdf

    def _guardar_resultados_geoespaciales(
        self,
        resultados: Dict[str, gpd.GeoDataFrame],
        output_dir: Path,
        año: str,
        seccion: str,
    ) -> None:
        """Guarda resultados geoespaciales en formato GeoJSON"""
        for categoria, gdf in resultados.items():
            if not gdf.empty:
                gdf_limpio = self._limpiar_columnas_geometria(gdf)
                nombre_archivo = f"{categoria.replace(' ', '_').lower()}_{año}_seccion_{seccion}.geojson"
                gdf_limpio.to_file(
                    output_dir / nombre_archivo, driver="GeoJSON")

    def generar_consolidado(self, todos_resultados: List[ResultadosAnalisis]) -> pd.DataFrame:
        """Genera consolidado de métricas con criterios de decisión optimizado"""

        decisiones = {0: "Fiable", 1: "Condicional"}
        umbrales = (self.config.umbral_completeness,
                    self.config.umbral_sci, self.config.umbral_accuracy)

        consolidado = []

        for resultado in todos_resultados:
            m = resultado.metricas  # Alias para reducir repetición
            completeness, accuracy, sci = m["Completeness (%)"], m["Accuracy (%)"], m["SCI"] * 100

            # Contar fallos y determinar decisión en una línea
            fallos = sum(v < u for v, u in zip(
                [completeness, sci, accuracy], umbrales))
            decision = decisiones.get(fallos, "No fiable")

            consolidado.append({
                "Sección": resultado.seccion,
                "Completeness (%)": round(completeness, 2),
                "Accuracy (%)": round(accuracy, 2),
                "SCI": round(sci, 2),
                "Partial Match": m["Partial Match (Buffer) (n)"],
                "Omission": m["Omission (n)"],
                "Commission": m["Commission (n)"],
                "Decisión": decision,
            })

        return pd.DataFrame(consolidado)

    def generar_tabla_resumen_global(self, todos_resultados: List[ResultadosAnalisis]) -> pd.DataFrame:
        """Genera tabla resumen de métricas globales como conclusión optimizada"""
        if not todos_resultados:
            return pd.DataFrame()

        # Extraer y calcular estadísticas en una sola pasada
        metricas = ["Completeness (%)", "Accuracy (%)", "SCI"]
        nombres = ["Completitud", "Precisión", "SCI"]

        def _calcular_stats(valores):
            arr = np.array(valores)
            return {stat: func(arr) for stat, func in
                    zip(['mean', 'min', 'max', 'std'], [np.mean, np.min, np.max, np.std])}

        # Procesar todas las métricas de una vez
        stats = {}
        for i, metrica in enumerate(metricas):
            valores = [r.metricas[metrica] *
                       (100 if metrica == "SCI" else 1) for r in todos_resultados]
            stats[nombres[i]] = _calcular_stats(valores)

        # Construir DataFrame directamente con comprensión de diccionario
        df_resumen = pd.DataFrame([{
            "Métrica": nombres,
            **{col: [round(stats[nombre][stat], 2) for nombre in nombres]
               for col, stat in [("Valor (%)", 'mean'), ("Min (%)", 'min'),
                                 ("Max (%)", 'max'), ("Desviación Std", 'std')]}
        }])

        # Guardar Excel con ancho automático
        excel_path = "resumen_global_metricas.xlsx"
        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            df_resumen.to_excel(
                writer, sheet_name="Resumen_Global", index=False)
            ws = writer.sheets["Resumen_Global"]
            for col in ws.columns:
                max_len = max(len(str(cell.value))
                              if cell.value else 0 for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(
                    max_len + 2, 50)

        return df_resumen

    def calcular_estadisticas_globales(self, todos_resultados: List[ResultadosAnalisis]) -> dict:
        """
        Calcula estadísticas descriptivas completas para todas las métricas clave y las guarda en Excel.
        """
        if not todos_resultados:
            logger.warning(
                "No hay resultados para calcular estadísticas globales")
            return {}

        metricas_nombres = [
            "Completeness (%)", "Accuracy (%)", "SCI", "Partial Match (Buffer) (n)",
            "Omission (n)", "Commission (n)", "Longitud oficial (m)", "Longitud representada (m)"
        ]


    def detectar_fase_actual(self) -> str:
        """
        Detecta automáticamente en qué fase se encuentra el análisis optimizado
        Returns: 'sin_iniciar', 'diagnostico_completo', 'validacion_pendiente', 'post_validacion_completo'
        """
        output_dir = self.config.output_dir

        # Verificar si existe consolidado preliminar (Fase 1 completada)

        if not (output_dir / "metricas_automaticas_prevalidacion.xlsx").exists():
            return "sin_iniciar"
        # Buscar archivos de validación combinando patrones

        patrones_validacion = [
            "**/validacion_visual_unificada_*.xlsx",
            "**/dummy_validacion_visual_unificada_*.xlsx",
        ]
        todos_archivos_validacion = []
        for patron in patrones_validacion:
            todos_archivos_validacion.extend(output_dir.glob(patron))
        if not todos_archivos_validacion:
            return "diagnostico_completo"
        # Verificar validaciones completadas de forma optimizada

        archivos_validados = 0
        columna_validacion = "¿Es realmente un carril bici?"

        for archivo in todos_archivos_validacion:
            try:
                df = pd.read_excel(archivo)
                if columna_validacion not in df.columns:
                    continue
                # Contar respuestas válidas de forma optimizada

                respuestas = df[columna_validacion].fillna(
                    "").astype(str).str.strip()
                validadas = respuestas.str.startswith(("✅", "❌")).sum()

                if validadas > 0:
                    archivos_validados += 1
                    logger.info(
                        f"Archivo validado detectado: {archivo.name} ({validadas} validaciones)"
                    )
            except Exception as e:
                logger.warning(f"Error leyendo archivo {archivo}: {e}")
                continue
        # Determinar fase final

        reporte_final = output_dir / "media_resumen_metricas_validados.xlsx"

        if reporte_final.exists():
            return "post_validacion_completo"
        elif archivos_validados > 0:
            logger.info(
                f"Se detectaron {archivos_validados} archivos con validaciones completadas"
            )
            return "validacion_parcial"
        else:
            logger.info(
                "Archivos de validación encontrados pero sin completar")
            return "validacion_pendiente"

    def ejecutar_analisis_completo(self, fase: str = "auto") -> None:
        """
        Ejecuta el análisis en dos fases:
        - fase="auto": Detecta automáticamente la fase actual
        - fase="diagnostico": Genera archivos de validación (PAUSA AQUÍ)
        - fase="post_validacion": Procesa archivos validados y genera métricas finales
        - fase="forzar_diagnostico": Fuerza re-ejecución de diagnóstico
        """

        def _confirmar_accion(mensaje):
            """Función auxiliar para confirmaciones"""
            return input(f"{mensaje} [s/N]: ").lower() in ["s", "si", "y", "yes"]

        # Mapeo de estados y acciones
        acciones_estado = {
            "sin_iniciar": ("➡️ Ejecutando Fase 1: Diagnóstico inicial", "diagnostico"),
            "diagnostico_completo": ("✅ Fase 1 ya completada\n❌ No se detectaron validaciones manuales\n📋 Proceda a completar los archivos de validación o use fase='post_validacion' para procesar validaciones existentes", None),
            "validacion_pendiente": ("✅ Fase 1 ya completada\n⏳ Archivos de validación encontrados pero sin completar\n📋 Complete la validación manual y luego use fase='post_validacion'", None),
            "validacion_parcial": ("✅ Fase 1 completada\n⚠️ Validaciones parcialmente completadas detectadas", "confirmar_post"),
            "post_validacion_completo": ("✅ Análisis completo ya finalizado\n📊 Revise los reportes en la carpeta de resultados", "confirmar_rerun")
        }

        if fase == "auto":
            estado_actual = self.detectar_fase_actual()
            logger.info(f"🔍 Estado detectado: {estado_actual}")

            mensaje, accion = acciones_estado.get(
                estado_actual, ("Estado desconocido", None))
            logger.info(mensaje)

            if accion == "diagnostico":
                fase = "diagnostico"
            elif accion == "confirmar_post":
                fase = "post_validacion" if _confirmar_accion(
                    "¿Proceder con Fase 2 (post-validación)?") else None
            elif accion == "confirmar_rerun":
                fase = "post_validacion" if _confirmar_accion(
                    "¿Re-ejecutar Fase 2 (post-validación)?") else None
            else:
                return  # Sin acción, terminar

            if fase is None:
                logger.info("❌ Operación cancelada")
                return

        # Ejecutar fase determinada
        if fase in ["diagnostico", "forzar_diagnostico"]:
            if fase == "forzar_diagnostico":
                logger.warning(
                    "⚠️ FORZANDO re-ejecución de diagnóstico - se sobreescribirán archivos existentes")

            self._ejecutar_diagnostico_inicial()
            self.resumen_registros_a_validar()

            logger.info("\n" + "=" * 60)
            logger.info("FASE 1 COMPLETADA - DIAGNÓSTICO INICIAL")
            logger.info("=" * 60)
            logger.info(
                "SIGUIENTE PASO: Completar validación manual en archivos:")
            logger.info("validacion_visual_unificada_*.xlsx")
            logger.info(
                "Luego ejecutar con fase='post_validacion' o sin argumentos (auto)")
            logger.info("=" * 60)

        elif fase == "post_validacion":
            self._ejecutar_post_validacion()

        else:
            logger.error(
                "Fase no válida. Use 'auto', 'diagnostico', 'post_validacion' o 'forzar_diagnostico'")

    def _ejecutar_diagnostico_inicial(self) -> None:
        """Ejecuta la fase de diagnóstico inicial - genera archivos para validación manual"""
        logger.info("INICIANDO FASE 1: DIAGNÓSTICO INICIAL\n" + "=" * 50)

        # Obtener y clasificar secciones
        secciones = self.obtener_secciones_unicas()
        secciones_con_oficial, secciones_sin_oficial = self.clasificar_secciones(
            secciones)

        # Procesar secciones CON geometría oficial
        todos_resultados = self._procesar_secciones_con_oficial(
            secciones_con_oficial)

        # Procesar secciones SIN geometría oficial
        todos_resultados.extend(
            self._procesar_secciones_sin_oficial(secciones_sin_oficial))

        # Identificar secciones sin oficial con commission
        self.identificar_secciones_sin_oficial_con_commission(
            secciones_sin_oficial)

        # Generar outputs finales
        if todos_resultados:
            self._generar_outputs_consolidados(todos_resultados)

        logger.info("\nFASE 1 COMPLETADA - Archivos de validación generados")
        logger.info(
            "Proceda a completar la validación manual antes de ejecutar la Fase 2")

    def _procesar_secciones_con_oficial(self, secciones_con_oficial: list) -> list:
        """Procesa secciones con geometría oficial usando ThreadPoolExecutor"""
        logger.info(
            f"\nProcesando {len(secciones_con_oficial)} secciones CON geometría oficial...")

        todos_resultados = []
        with ThreadPoolExecutor(max_workers=self.config.max_workers) as executor:
            futures = {executor.submit(self.procesar_seccion, seccion): seccion
                       for seccion in secciones_con_oficial}

            for future in as_completed(futures):
                try:
                    todos_resultados.extend(future.result())
                except Exception as e:
                    logger.error(
                        f"Error procesando sección {futures[future]}: {e}")

        return todos_resultados

    def _procesar_secciones_sin_oficial(self, secciones_sin_oficial: list) -> list:
        """Procesa secciones sin geometría oficial"""
        logger.info(
            f"\nProcesando {len(secciones_sin_oficial)} secciones SIN geometría oficial...")

        todos_resultados = []
        for seccion in secciones_sin_oficial:
            try:
                todos_resultados.extend(
                    self.procesar_seccion_sin_oficial(seccion))
            except Exception as e:
                logger.error(
                    f"Error procesando sección sin oficial {seccion}: {e}")

        return todos_resultados

    def _generar_outputs_consolidados(self, todos_resultados: list) -> None:
        """Genera consolidado y outputs finales"""
        consolidado = self.generar_consolidado(todos_resultados)
        self.generar_graficos_metricas_automaticas(consolidado)
        self.generar_mapa_fiabilidad_final(consolidado)
        self._exportar_excel_optimizado(consolidado)

        logger.info(
            f"Consolidado preliminar generado: {len(consolidado)} secciones procesadas")

    def _exportar_excel_optimizado(self, consolidado) -> None:
        """Exporta el consolidado a Excel con columnas auto-ajustadas"""
        archivo_excel = self.config.output_dir / \
            "metricas_automaticas_prevalidacion.xlsx"

        with pd.ExcelWriter(archivo_excel, engine="openpyxl") as writer:
            consolidado.to_excel(writer, index=False, sheet_name="Consolidado")

            # Auto-ajustar columnas
            worksheet = writer.sheets["Consolidado"]
            for column in worksheet.columns:
                max_length = max((len(str(cell.value))
                                 for cell in column if cell.value), default=0)
                worksheet.column_dimensions[column[0].column_letter].width = min(
                    max_length + 2, 50)

    def _ejecutar_post_validacion(self) -> None:
        """Ejecuta la fase post-validación manual - genera métricas finales"""
        logger.info("INICIANDO FASE 2: POST-VALIDACIÓN MANUAL\n" + "=" * 50)

        # Buscar todos los archivos de validación (incluir dummy)
        todos_archivos_validacion = (
            list(self.config.output_dir.glob("**/validacion_visual_unificada_*.xlsx")) +
            list(self.config.output_dir.glob(
                "**/dummy_validacion_visual_unificada_*.xlsx"))
        )

        if not todos_archivos_validacion:
            logger.error(
                "No se encontraron archivos de validación completados")
            logger.error(
                "Asegúrese de haber completado la validación manual en la Fase 1")
            return

        logger.info(
            f"Encontrados {len(todos_archivos_validacion)} archivos de validación")

        # Procesar archivos validados
        resultados_validados = []
        metricas_corregidas = []

        for archivo in todos_archivos_validacion:
            try:
                resultado = self._procesar_archivo_validacion(archivo)
                if resultado:
                    metricas_corregidas.append(resultado['metricas'])
                    resultados_validados.append(resultado['resultado_obj'])
                    logger.info(
                        f"Procesada validación: {resultado['seccion']} - {resultado['año']}")
            except Exception as e:
                logger.error(f"Error procesando archivo {archivo}: {e}")

        # Generar reportes finales
        if metricas_corregidas:
            self._generar_reportes_finales(metricas_corregidas)
            logger.info(
                f"Procesadas {len(metricas_corregidas)} secciones con validaciones completadas")

            # Generar estadísticas y mapas
            self.generar_mapas_tematicos(resultados_validados)
            self.calcular_estadisticas_globales(resultados_validados)
        else:
            logger.error(
                "No se encontraron validaciones completadas para procesar")
            return

        logger.info("\nFASE 2 COMPLETADA - Análisis post-validación finalizado")

    def _procesar_archivo_validacion(self, archivo) -> dict:
        """Procesa un archivo individual de validación"""
        # Extraer información del nombre del archivo
        nombre = archivo.name.replace("dummy_", "").replace(
            "validacion_visual_unificada_", "").replace(".xlsx", "")
        partes = nombre.split("_")
        año, seccion = partes[0], partes[-1]

        # Leer y validar archivo
        df_validacion = pd.read_excel(archivo)

        if "¿Es realmente un carril bici?" not in df_validacion.columns:
            logger.warning(
                f"Archivo {archivo.name} no tiene columna de validación - saltando")
            return None

        # Verificar validaciones completadas
        respuestas = df_validacion["¿Es realmente un carril bici?"].fillna(
            "").astype(str)
        validadas = sum(1 for r in respuestas if r.strip()
                        and (r.startswith("✅") or r.startswith("❌")))

        if validadas == 0:
            logger.warning(
                f"Archivo {archivo.name} no tiene validaciones completadas - saltando")
            return None

        logger.info(f"Procesando: {archivo.name} ({validadas} validaciones)")

        # Procesar validación manual
        metricas_corregidas_seccion = self._procesar_validacion_manual(
            df_validacion, seccion, año)
        if not metricas_corregidas_seccion:
            return None

        # Obtener longitudes de métricas previas
        longitud_oficial, longitud_representada = self._obtener_longitudes_previas(
            seccion, año)

        # Crear objeto ResultadosAnalisis
        resultado_obj = self._crear_resultado_analisis(
            seccion, año, metricas_corregidas_seccion, longitud_oficial, longitud_representada
        )

        return {
            'metricas': metricas_corregidas_seccion,
            'resultado_obj': resultado_obj,
            'seccion': seccion,
            'año': año
        }

    def _obtener_longitudes_previas(self, seccion: str, año: str) -> tuple:
        """Obtiene las longitudes oficial y representada de métricas previas"""
        metricas_path = self.config.output_dir / \
            f"seccion_{seccion}" / f"metricas_{seccion}_{año}.xlsx"
        longitud_oficial = longitud_representada = 0

        if metricas_path.exists():
            df_metricas = pd.read_excel(metricas_path)
            if "Longitud oficial (m)" in df_metricas.columns:
                longitud_oficial = df_metricas["Longitud oficial (m)"].iloc[0]
            if "Longitud representada (m)" in df_metricas.columns:
                longitud_representada = df_metricas["Longitud representada (m)"].iloc[0]

        return longitud_oficial, longitud_representada

    def _crear_resultado_analisis(self, seccion: str, año: str, metricas: dict,
                                  longitud_oficial: float, longitud_representada: float) -> ResultadosAnalisis:
        """Crea un objeto ResultadosAnalisis con las métricas corregidas"""
        # GeoDataFrames vacíos
        empty_gdf = gpd.GeoDataFrame(geometry=[])

        return ResultadosAnalisis(
            seccion=seccion,
            año=año,
            partial_match=empty_gdf,
            omission=empty_gdf,
            commission=empty_gdf,
            metricas={
                "Completeness (%)": metricas.get("Recall (%)", 0),
                "Accuracy (%)": metricas.get("Precision (%)", 0),
                "SCI": metricas.get("F1_Score", 0) / 100,
                "Partial Match (Buffer) (n)": metricas.get("Commission_Confirmados", 0),
                "Omission (n)": metricas.get("Omission_Confirmados", 0),
                "Commission (n)": metricas.get("Commission_Falsos", 0),
                "Longitud oficial (m)": longitud_oficial,
                "Longitud representada (m)": longitud_representada,
            }
        )

    def _procesar_validacion_manual(self, df_validacion: pd.DataFrame, seccion: str, año: str) -> Dict[str, any]:
        """Procesa los resultados de validación manual y calcula métricas corregidas"""

        if "¿Es realmente un carril bici?" not in df_validacion.columns:
            logger.warning(
                f"Columna de validación no encontrada en sección {seccion}")
            return None

        # Contar validaciones por categoría
        conteos = {"commission_confirmados": 0, "commission_falsos": 0,
                   "omission_confirmados": 0, "omission_falsos": 0}

        for _, row in df_validacion.iterrows():
            categoria = row.get("Categoria", "")
            validacion = str(
                row.get("¿Es realmente un carril bici?", "")).strip()

            if categoria == "Commission":
                conteos["commission_confirmados" if validacion.startswith(
                    "✅") else "commission_falsos"] += validacion.startswith(("✅", "❌"))
            elif categoria == "Omission":
                conteos["omission_confirmados" if validacion.startswith(
                    "✅") else "omission_falsos"] += validacion.startswith(("✅", "❌"))

        # Calcular métricas corregidas
        tp, fp, fn = conteos["commission_confirmados"], conteos["commission_falsos"], conteos["omission_confirmados"]

        precision = (tp / (tp + fp) * 100) if (tp + fp) > 0 else 0
        recall = (tp / (tp + fn) * 100) if (tp + fn) > 0 else 0
        f1_score = (2 * precision * recall / (precision + recall)
                    ) if (precision + recall) > 0 else 0

        return {
            "Seccion": seccion, "Año": año,
            "Commission_Confirmados": conteos["commission_confirmados"],
            "Commission_Falsos": conteos["commission_falsos"],
            "Omission_Confirmados": conteos["omission_confirmados"],
            "Omission_Falsos": conteos["omission_falsos"],
            "True_Positives": tp, "False_Positives": fp, "False_Negatives": fn,
            "Precision (%)": round(precision, 2),
            "Recall (%)": round(recall, 2),
            "F1_Score": round(f1_score, 2),
            "Total_Validados": len(df_validacion),
        }

    def generar_graficos_metricas_automaticas(self, df_consolidado: pd.DataFrame) -> None:
        """Genera histogramas y boxplots de métricas automáticas (Fase 1) para detectar variabilidad y outliers."""

        if df_consolidado.empty:
            print("❌ No hay datos para generar gráficos automáticos.")
            return

        metricas = ["Completeness (%)", "Accuracy (%)", "SCI"]
        fig, axes = plt.subplots(len(metricas), 2, figsize=(12, 10))
        fig.suptitle("Distribución de métricas automáticas (Fase 1)",
                     fontsize=16, fontweight="bold")

        for i, metrica in enumerate(metricas):
            valores = df_consolidado[metrica].dropna()

            # Histograma
            self._crear_histograma(axes[i, 0], valores, metrica)

            # Boxplot
            self._crear_boxplot(axes[i, 1], valores, metrica)

        plt.tight_layout(rect=[0, 0.03, 1, 0.95])
        output_path = self.config.output_dir / "graficos_metricas_automaticas.png"
        plt.savefig(output_path, dpi=300, bbox_inches="tight")
        plt.close()
        print(f"✅ Gráficos automáticos guardados en: {output_path}")

    def _crear_histograma(self, ax, valores, metrica: str) -> None:
        """Crea un histograma para una métrica específica"""
        ax.hist(valores, bins=15, color="cornflowerblue",
                edgecolor="black", alpha=0.7)
        ax.axvline(valores.mean(), color="red", linestyle="--",
                   label=f"Media: {valores.mean():.1f}")
        ax.set_title(f"{metrica} - Histograma")
        ax.legend()

    def _crear_boxplot(self, ax, valores, metrica: str) -> None:
        """Crea un boxplot para una métrica específica"""
        ax.boxplot(valores, vert=False, patch_artist=True,
                   boxprops=dict(facecolor="lightgray"))
        ax.set_title(f"{metrica} - Boxplot")

    def _generar_reportes_finales(self, metricas_corregidas: List[Dict]) -> None:
        """Genera reportes finales con métricas corregidas post-validación"""

        df_final = pd.DataFrame(metricas_corregidas)

        # Generar archivos Excel
        self._generar_excel_metricas(df_final)
        self.generar_excel_resumen_metricas_post_validacion()

        # Generar contenido adicional
        self.generar_mapas_tematicos_post_validacion()

        # Log de archivos generados
        archivos_generados = [
            "metricas_validadas_post_GSV.xlsx",
            "media_resumen_metricas_validados.xlsx",
            "graficos_resultados_finales.png"
        ]

        logger.info("Reportes finales generados:")
        for archivo in archivos_generados:
            logger.info(f"- {archivo}")

    def _generar_excel_metricas(self, df_final: pd.DataFrame) -> None:
        """Genera archivo Excel con métricas detalladas"""
        metricas_path = self.config.output_dir / "metricas_validadas_post_GSV.xlsx"
        with pd.ExcelWriter(metricas_path, engine="openpyxl") as writer:
            df_final.to_excel(writer, sheet_name="Métricas", index=False)
            self._autoajustar_columnas(writer, df_final, "Métricas")

    def generar_excel_resumen_metricas_post_validacion(self):
        """
        Genera resumen único de métricas validadas por tipo de sección:
        - CON infraestructura: Accuracy, Completeness y SCI
        - SIN infraestructura: Solo Precision (%)
        - GLOBAL: Todas las métricas combinadas
        Además, exporta dos archivos separados por tipo incluyendo una fila TOTAL.
        """
        # === Cargar archivos necesarios ===
        path_post = self.config.output_dir / "metricas_validadas_post_GSV.xlsx"
        path_pre = self.config.output_dir / "metricas_automaticas_prevalidacion.xlsx"
        path_con = self.config.output_dir / "secciones_con_oficial.xlsx"
        path_sin = self.config.output_dir / "secciones_sin_oficial.xlsx"

        df_post = pd.read_excel(path_post)
        df_post["Seccion"] = df_post["Seccion"].astype(str).str.zfill(11)

        df_pre = pd.read_excel(path_pre)
        df_pre["Seccion"] = df_pre["Sección"].astype(str).str.zfill(11)
        df_pre = df_pre.rename(columns={
            "Partial Match (Buffer) (m)": "Partial Match",
            "Longitud oficial (m)": "Longitud_Oficial_m",
            "SCI": "SCI"
        })

        df_con = pd.read_excel(path_con)
        df_sin = pd.read_excel(path_sin)
        set_con = set(df_con["MUNDISSEC"].astype(str).str.zfill(11))
        set_sin = set(df_sin["MUNDISSEC"].astype(str).str.zfill(11))

        def clasificar(seccion):
            if seccion in set_con:
                return "CON infraestructura oficial"
            elif seccion in set_sin:
                return "SIN infraestructura oficial"
            else:
                return "desconocido"

        df_post["Tipo_Seccion"] = df_post["Seccion"].apply(clasificar)
        df_pre["Tipo_Seccion"] = df_pre["Seccion"].apply(clasificar)

        resumen = []

        # === CON infraestructura ===
        df_con_post = df_post[df_post["Tipo_Seccion"] == "CON infraestructura oficial"].copy()
        df_con_pre = df_pre[df_pre["Tipo_Seccion"] == "CON infraestructura oficial"].copy()

        tp_partial = df_con_pre["Partial Match"].sum()
        tp_commissions = df_con_post["True_Positives"].sum()
        tp_total = tp_partial + tp_commissions
        fp = df_con_post["False_Positives"].sum()
        fn = df_con_post["False_Negatives"].sum()
        sci = df_con_pre["SCI"].mean()

        accuracy = tp_total / (tp_total + fp + fn) * 100 if (tp_total + fp + fn) > 0 else None
        completeness = tp_total / (tp_total + fn) * 100 if (tp_total + fn) > 0 else None

        resumen.append({
            "Tipo de Sección": "CON infraestructura oficial",
            "Accuracy (%)": round(accuracy, 2) if accuracy is not None else None,
            "Completeness (%)": round(completeness, 2) if completeness is not None else None,
            "SCI (%)": round(sci, 2) if sci is not None else None,
            "Secciones Analizadas": len(df_con_post),
            "Total Validados": df_con_post["Total_Validados"].sum()
        })

        # Exportar métricas CON infraestructura con fila TOTAL
        df_con_post = df_con_post.rename(columns={
            "True_Positives": "TP visuales",
            "False_Positives": "FP visuales",
            "False_Negatives": "FN visuales"
        })
        df_con_pre = df_con_pre.rename(columns={
            "Partial Match": "Partial Match (m)",
            "SCI": "SCI (%)"
        })
        df_con_merged = df_con_post.merge(df_con_pre[["Seccion", "Partial Match (m)", "SCI (%)"]], on="Seccion", how="left")
        df_con_merged["TP total"] = df_con_merged["TP visuales"] + df_con_merged["Partial Match (m)"]
        df_con_merged["Accuracy (%)"] = df_con_merged.apply(
            lambda row: round(row["TP total"] / (row["TP total"] + row["FP visuales"] + row["FN visuales"]) * 100, 2)
            if (row["TP total"] + row["FP visuales"] + row["FN visuales"]) > 0 else None, axis=1)
        df_con_merged["Completeness (%)"] = df_con_merged.apply(
            lambda row: round(row["TP total"] / (row["TP total"] + row["FN visuales"]) * 100, 2)
            if (row["TP total"] + row["FN visuales"]) > 0 else None, axis=1)

        columnas_finales_con = [
            "Seccion", "TP visuales", "Partial Match (m)", "TP total",
            "FP visuales", "FN visuales", "Accuracy (%)", "Completeness (%)", "SCI (%)"
        ]

        # Añadir fila TOTAL
        fila_total_con = {
            "Seccion": "TOTAL",
            "TP visuales": df_con_merged["TP visuales"].sum(),
            "Partial Match (m)": df_con_merged["Partial Match (m)"].sum(),
            "TP total": df_con_merged["TP total"].sum(),
            "FP visuales": df_con_merged["FP visuales"].sum(),
            "FN visuales": df_con_merged["FN visuales"].sum(),
            "Accuracy (%)": round(df_con_merged["TP total"].sum() / (
                df_con_merged["TP total"].sum() + df_con_merged["FP visuales"].sum() + df_con_merged["FN visuales"].sum()) * 100, 2)
            if (df_con_merged["TP total"].sum() + df_con_merged["FP visuales"].sum() + df_con_merged["FN visuales"].sum()) > 0 else None,
            "Completeness (%)": round(df_con_merged["TP total"].sum() / (
                df_con_merged["TP total"].sum() + df_con_merged["FN visuales"].sum()) * 100, 2)
            if (df_con_merged["TP total"].sum() + df_con_merged["FN visuales"].sum()) > 0 else None,
            "SCI (%)": round(df_con_merged["SCI (%)"].mean(), 2)
        }

        df_con_export = pd.concat([df_con_merged[columnas_finales_con], pd.DataFrame([fila_total_con])], ignore_index=True)
        df_con_export.to_excel(self.config.output_dir / "metricas_con_infraestructura_validacion.xlsx", index=False)

        # === SIN infraestructura ===
        df_sin_post = df_post[df_post["Tipo_Seccion"] == "SIN infraestructura oficial"].copy()
        tp_sin = df_sin_post["True_Positives"].sum()
        fp_sin = df_sin_post["False_Positives"].sum()
        precision_sin = tp_sin / (tp_sin + fp_sin) * 100 if (tp_sin + fp_sin) > 0 else None

        resumen.append({
            "Tipo de Sección": "SIN infraestructura oficial",
            "Precision (%)": round(precision_sin, 2) if precision_sin is not None else None,
            "Secciones Analizadas": len(df_sin_post),
            "Total Validados": df_sin_post["Total_Validados"].sum()
        })

        # Exportar métricas SIN infraestructura con fila TOTAL
        df_sin_export = df_sin_post.rename(columns={
            "True_Positives": "TP",
            "False_Positives": "FP",
            "Total_Validados": "Total de tramos validados"
        })
        df_sin_export["Precision (%)"] = df_sin_export.apply(
            lambda row: round(row["TP"] / (row["TP"] + row["FP"]) * 100, 2)
            if (row["TP"] + row["FP"]) > 0 else None, axis=1)

        df_sin_export = df_sin_export[["Seccion", "Año", "TP", "FP", "Precision (%)", "Total de tramos validados"]]

        fila_total_sin = {
            "Seccion": "TOTAL",
            "Año": "",
            "TP": df_sin_export["TP"].sum(),
            "FP": df_sin_export["FP"].sum(),
            "Precision (%)": round(df_sin_export["TP"].sum() / (
                df_sin_export["TP"].sum() + df_sin_export["FP"].sum()) * 100, 2)
            if (df_sin_export["TP"].sum() + df_sin_export["FP"].sum()) > 0 else None,
            "Total de tramos validados": df_sin_export["Total de tramos validados"].sum()
        }

        df_sin_export = pd.concat([df_sin_export, pd.DataFrame([fila_total_sin])], ignore_index=True)
        df_sin_export.to_excel(self.config.output_dir / "metricas_sin_infraestructura_validacion.xlsx", index=False)

        # === GLOBAL ===
        df_global_post = df_post[df_post["Tipo_Seccion"] != "desconocido"]
        df_global_pre = df_pre[df_pre["Tipo_Seccion"] == "CON infraestructura oficial"]

        tp_global = df_global_post["True_Positives"].sum() + df_global_pre["Partial Match"].sum()
        fp_global = df_global_post["False_Positives"].sum()
        fn_global = df_global_post["False_Negatives"].sum()

        accuracy_global = tp_global / (tp_global + fp_global + fn_global) * 100 if (tp_global + fp_global + fn_global) > 0 else None
        completeness_global = tp_global / (tp_global + fn_global) * 100 if (tp_global + fn_global) > 0 else None
        sci_global = df_global_pre["SCI"].mean()
        precision_global = tp_global / (tp_global + fp_global) * 100 if (tp_global + fp_global) > 0 else None

        resumen.append({
            "Tipo de Sección": "GLOBAL",
            "Accuracy (%)": round(accuracy_global, 2) if accuracy_global is not None else None,
            "Completeness (%)": round(completeness_global, 2) if completeness_global is not None else None,
            "SCI (%)": round(sci_global, 2) if sci_global is not None else None,
            "Precision (%)": round(precision_global, 2) if precision_global is not None else None,
            "Secciones Analizadas": len(df_global_post),
            "Total Validados": df_global_post["Total_Validados"].sum()
        })

        # Exportar resumen global
        df_resumen = pd.DataFrame(resumen)
        output_resumen = self.config.output_dir / "resumen_metricas_validadas_por_tipo.xlsx"
        df_resumen.to_excel(output_resumen, index=False)

        print(f"✅ Resumen exportado correctamente a: {output_resumen}")
        print("📁 Archivos individuales exportados:")
        print("   - metricas_con_infraestructura_validacion.xlsx")
        print("   - metricas_sin_infraestructura_validacion.xlsx")
        # === Exportar resumen completo a .txt (PRE + POST) ===
        resumen_txt_path = self.config.output_dir / "resumen_metricas_pre_y_post_validacion.txt"
        with open(resumen_txt_path, "w", encoding="utf-8") as f:
            f.write("📊 RESUMEN DE MÉTRICAS POST-VALIDACIÓN POR TIPO DE SECCIÓN\n")
            f.write("=" * 70 + "\n\n")
            for row in resumen:
                f.write(f"🔸 {row['Tipo de Sección']}\n")
                f.write("-" * 40 + "\n")
                for key, value in row.items():
                    if key != "Tipo de Sección":
                        f.write(f"{key}: {value}\n")
                f.write("\n")

            # === Añadir bloque PREVALIDACIÓN ===
            f.write("\n📊 RESUMEN DE MÉTRICAS AUTOMÁTICAS PREVALIDACIÓN (solo secciones con infraestructura)\n")
            f.write("=" * 70 + "\n\n")

            try:
                path_pre = self.config.output_dir / "metricas_automaticas_prevalidacion.xlsx"
                path_con = self.config.output_dir / "secciones_con_oficial.xlsx"

                df_pre = pd.read_excel(path_pre)
                df_pre["Seccion"] = df_pre["Sección"].astype(str).str.zfill(11)

                set_con = set(pd.read_excel(path_con)["MUNDISSEC"].astype(str).str.zfill(11))
                df_filtrado = df_pre[df_pre["Seccion"].isin(set_con)].copy()

                suma_ponderada_completeness = 0.0
                suma_ponderada_accuracy = 0.0
                suma_ponderada_sci = 0.0
                total_longitud = 0.0
                secciones_validas = 0

                for _, row in df_filtrado.iterrows():
                    seccion = row["Seccion"]
                    año = str(row.get("Año", "2024"))  # usa el año por defecto si no está
                    completeness = row.get("Completeness (%)")
                    accuracy = row.get("Accuracy (%)")
                    sci = row.get("SCI", row.get("SCI (%)"))

                    if pd.isna(completeness) or pd.isna(accuracy) or pd.isna(sci):
                        continue

                    longitud_oficial, _ = self._obtener_longitudes_previas(seccion, año)
                    if longitud_oficial <= 0:
                        continue

                    suma_ponderada_completeness += completeness * longitud_oficial
                    suma_ponderada_accuracy += accuracy * longitud_oficial
                    suma_ponderada_sci += sci * longitud_oficial
                    total_longitud += longitud_oficial
                    secciones_validas += 1

                if total_longitud > 0:
                    promedio_completeness = suma_ponderada_completeness / total_longitud
                    promedio_accuracy = suma_ponderada_accuracy / total_longitud
                    promedio_sci = suma_ponderada_sci / total_longitud
                else:
                    promedio_completeness = promedio_accuracy = promedio_sci = 0

                f.write(f"Secciones con infraestructura válidas: {secciones_validas}\n")
                f.write(f"Longitud oficial total (m): {round(total_longitud, 2)}\n\n")
                f.write(f"✔ Completeness promedio ponderado: {promedio_completeness:.2f}%\n")
                f.write(f"✔ Accuracy promedio ponderado:     {promedio_accuracy:.2f}%\n")
                f.write(f"✔ SCI promedio ponderado:          {promedio_sci:.2f}%\n")

            except Exception as e:
                f.write("⚠️ No se pudo calcular el resumen de prevalidación.\n")
                f.write(f"Error: {str(e)}\n")




    def _autoajustar_columnas(self, writer, df: pd.DataFrame, sheet_name: str) -> None:
        """Autoajusta el ancho de las columnas en Excel"""
        worksheet = writer.sheets[sheet_name]
        for idx, col in enumerate(df.columns):
            max_len = max(
                len(str(col)),
                df[col].astype(str).str.len().max() if not df.empty else 0
            )
            worksheet.column_dimensions[worksheet.cell(
                1, idx + 1).column_letter].width = min(max_len + 2, 50)


    def generar_mapas_tematicos(self, todos_resultados: List[ResultadosAnalisis]) -> None:
        """Genera mapas temáticos de completitud, accuracy y SCI por sección censal,
        además de mapas comparativos entre red oficial (SHP) y red OSM."""

        if not todos_resultados:
            logger.warning("No hay resultados para generar mapas temáticos")
            return

        # 1. GENERAR MAPAS TEMÁTICOS DE MÉTRICAS
        datos_mapa = [
            {
                "SECCIO": resultado.seccion,
                "Completeness": resultado.metricas.get("Completeness (%)", 0),
                "Accuracy": resultado.metricas.get("Accuracy (%)", 0),
                "SCI": resultado.metricas.get("SCI", 0) * 100,
                "geometry": seccion_geom.geometry.iloc[0],
            }
            for resultado in todos_resultados
            if not (seccion_geom := self.shp_secciones[
                self.shp_secciones["MUNDISSEC"] == resultado.seccion
            ]).empty
        ]

        if not datos_mapa:
            logger.warning(
                "No se pudieron preparar datos para mapas temáticos")
            return

        gdf_mapas = gpd.GeoDataFrame(datos_mapa, crs="EPSG:4326")

        # Generar mapas de métricas
        fig, axes = plt.subplots(1, 3, figsize=(18, 6))
        plot_config = {"legend": True, "cmap": "RdYlGn",
                       "edgecolor": "black", "linewidth": 0.5}

        for i, metrica in enumerate(["Completeness", "Accuracy", "SCI"]):
            gdf_mapas.plot(column=metrica, ax=axes[i], **plot_config)
            axes[i].set_title(f"{metrica} (%)", fontsize=14)
            axes[i].set_axis_off()
        
        plt.tight_layout()
        plt.savefig(self.config.output_dir /
                    "mapas_tematicos_metricas.png", dpi=300, bbox_inches="tight")
        plt.close()
        logger.info("Mapas temáticos generados: mapas_tematicos_metricas.png")

        # 2. GENERAR MAPAS COMPARATIVOS SHP vs OSM
        try:
            # Cargar datos oficiales
            try:
                shp_ciclovias = gpd.read_file(self.config.path_oficial)
                crs_referencia = shp_ciclovias.crs
                logger.info(
                    f"Cargados datos oficiales: {len(shp_ciclovias)} segmentos")
            except Exception as e:
                logger.warning(
                    f"No se pudieron cargar datos de ciclovías oficiales: {e}")
                return

            # Buscar y combinar archivos GeoJSON de OSM
            tipos_archivos = ["omission", "commission", "partial_match_(buffer)"]
            osm_gdfs_por_tipo = {tipo: [] for tipo in tipos_archivos}
            todos_osm_gdfs = []

            for año in self.config.años:
                for seccion_dir in self.config.output_dir.glob("seccion_*"):
                    if not seccion_dir.is_dir():
                        continue

                    mundissec = seccion_dir.name.replace("seccion_", "")
                    for tipo in tipos_archivos:
                        for archivo in seccion_dir.glob(f"{tipo}_{año}_seccion_{mundissec}.geojson"):
                            try:
                                gdf_osm = gpd.read_file(archivo)
                                if not gdf_osm.empty:
                                    if gdf_osm.crs != crs_referencia:
                                        gdf_osm = gdf_osm.to_crs(
                                            crs_referencia)
                                    gdf_osm["seccion"] = mundissec
                                    gdf_osm["año"] = año
                                    gdf_osm["tipo_analisis"] = tipo
                                    osm_gdfs_por_tipo[tipo].append(gdf_osm)
                                    todos_osm_gdfs.append(gdf_osm)
                            except Exception as e:
                                logger.warning(
                                    f"Error cargando {archivo}: {e}")
            if not todos_osm_gdfs:
                logger.warning(
                    "No se encontraron archivos GeoJSON de OSM en las carpetas de secciones")
                return

            gdf_osm_combined = pd.concat(todos_osm_gdfs, ignore_index=True)
            logger.info(
                f"OSM combinado: {len(gdf_osm_combined)} segmentos en CRS {gdf_osm_combined.crs}")

            # Generar mapas combinados por tipo
            mapas_combinados = {}
            for tipo, gdfs_list in osm_gdfs_por_tipo.items():
                if gdfs_list:
                    try:
                        gdf_combinado = pd.concat(gdfs_list, ignore_index=True)
                        mapas_combinados[tipo] = gdf_combinado
                        logger.info(
                            f"Mapa combinado {tipo}: {len(gdf_combinado)} segmentos totales")

                        archivo_salida = self.config.output_dir / \
                            f"mapa_combinado_{tipo}_{'-'.join(self.config.años)}.geojson"
                        gdf_combinado.to_file(archivo_salida, driver="GeoJSON")
                    except Exception as e:
                        logger.error(
                            f"Error combinando mapas tipo {tipo}: {e}")

            # Guardar mapa OSM completo
            archivo_general = self.config.output_dir / \
                f"mapa_osm_completo_{'-'.join(self.config.años)}.geojson"
            gdf_osm_combined.to_file(archivo_general, driver="GeoJSON")
            logger.info(
                f"Mapa OSM completo guardado: {archivo_general} - {len(gdf_osm_combined)} segmentos")

            # Cargar secciones y calcular bounds
            shp_secciones = None
            try:
                shp_secciones = gpd.read_file(self.config.path_secciones)
                if shp_secciones.crs != crs_referencia:
                    shp_secciones = shp_secciones.to_crs(crs_referencia)
            except Exception as e:
                logger.warning(f"No se pudieron cargar las secciones: {e}")

            bounds_shp, bounds_osm = shp_ciclovias.total_bounds, gdf_osm_combined.total_bounds

            if shp_secciones is not None:
                bounds_secciones = shp_secciones.total_bounds
                minx, miny = min(bounds_shp[0], bounds_osm[0], bounds_secciones[0]), min(
                    bounds_shp[1], bounds_osm[1], bounds_secciones[1])
                maxx, maxy = max(bounds_shp[2], bounds_osm[2], bounds_secciones[2]), max(
                    bounds_shp[3], bounds_osm[3], bounds_secciones[3])
                logger.info(
                    "Usando bounds de secciones para vista completa del área")
            else:
                minx, miny = min(bounds_shp[0], bounds_osm[0]), min(
                    bounds_shp[1], bounds_osm[1])
                maxx, maxy = max(bounds_shp[2], bounds_osm[2]), max(
                    bounds_shp[3], bounds_osm[3])
                margin_x, margin_y = (maxx - minx) * 0.1, (maxy - miny) * 0.1
                minx, miny, maxx, maxy = minx - margin_x, miny - \
                    margin_y, maxx + margin_x, maxy + margin_y
                logger.info("Usando bounds de ciclovías con margen adicional")

            # Función auxiliar para configurar mapas
            def configurar_mapa(ax, titulo, elementos_count):
                if shp_secciones is not None:
                    shp_secciones.plot(
                        ax=ax, color="lightgray", alpha=0.3, edgecolor="gray", linewidth=0.5)
                ax.set_xlim(minx, maxx)
                ax.set_ylim(miny, maxy)
                ax.set_title(titulo, fontsize=16, pad=20)
                ax.set_axis_off()
                ax.text(0.02, 0.98, f"Total segmentos: {elementos_count}", transform=ax.transAxes,
                        verticalalignment="top", bbox=dict(boxstyle="round", facecolor="white", alpha=0.8))

            # Crear mapas lado a lado
            fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(20, 10))

            configurar_mapa(
                ax1, "Red de Ciclovías Oficial (Ayuntamiento)", len(shp_ciclovias))
            shp_ciclovias.plot(ax=ax1, color="blue", linewidth=2, alpha=0.8)

            configurar_mapa(ax2, "Red de Ciclovías OpenStreetMap",
                            len(gdf_osm_combined))
            gdf_osm_combined.plot(ax=ax2, color="red",
                                  linewidth=1.5, alpha=0.8)
            
            plt.tight_layout()
            plt.savefig(self.config.output_dir /
                        "comparativa_redes_shp_vs_osm.png", dpi=300, bbox_inches="tight")
            plt.close()

            # Crear mapa superpuesto
            fig, ax = plt.subplots(1, 1, figsize=(15, 12))

            if shp_secciones is not None:
                shp_secciones.plot(ax=ax, color="lightgray",
                                   alpha=0.3, edgecolor="gray", linewidth=0.5)

            shp_ciclovias.plot(ax=ax, color="blue", linewidth=2,
                               alpha=0.7, label="Red Oficial")
            gdf_osm_combined.plot(ax=ax, color="red",
                                  linewidth=1, alpha=0.7, label="Red OSM")

            ax.set_xlim(minx, maxx)
            ax.set_ylim(miny, maxy)
            ax.set_title(
                "Comparativa: Red Oficial vs OpenStreetMap", fontsize=16, pad=20)
            ax.set_axis_off()
            ax.legend(loc="upper right", framealpha=0.9)

            info_text = f"Red Oficial: {len(shp_ciclovias)} segmentos\nRed OSM: {len(gdf_osm_combined)} segmentos"
            ax.text(0.02, 0.98, info_text, transform=ax.transAxes, verticalalignment="top",
                    bbox=dict(boxstyle="round", facecolor="white", alpha=0.8))

            plt.tight_layout()
            plt.savefig(self.config.output_dir /
                        "comparativa_redes_superpuestas.png", dpi=300, bbox_inches="tight")
            plt.close()

            # Logging final
            logger.info("Mapas generados exitosamente:")
            logger.info("- Mapas combinados por tipo de análisis (GeoJSON)")
            logger.info("- Mapa OSM completo (GeoJSON)")
            logger.info("- comparativa_redes_shp_vs_osm.png")
            logger.info("- comparativa_redes_superpuestas.png")
            logger.info(f"Red oficial: {len(shp_ciclovias)} segmentos")
            logger.info(f"Red OSM: {len(gdf_osm_combined)} segmentos")

            for tipo, gdf in mapas_combinados.items():
                logger.info(f"- {tipo}: {len(gdf)} segmentos")

        except Exception as e:
            logger.error(f"Error general en generación de mapas: {e}")
            return

    def generar_mapas_tematicos_post_validacion(self):  # TODO falta optimizar
        """
        Genera mapas temáticos de Precision, Recall y F1-Score por sección censal
        usando los resultados tras la validación manual (post-GSV).
        """

        # 1. Cargar métricas post-validación y geometría de secciones censales

        metricas_path = self.config.output_dir / "metricas_validadas_post_GSV.xlsx"
        secciones_path = self.config.path_secciones

        if not metricas_path.exists():
            print(f"❌ No se encuentra el archivo {metricas_path}")
            return
        df_metricas = pd.read_excel(metricas_path)
        gdf_secciones = gpd.read_file(secciones_path)
        gdf_secciones["MUNDISSEC"] = (
            gdf_secciones["MUNDISSEC"].astype(str).str.zfill(11)
        )

        # 2. Unir métricas a geometría (sólo las secciones validadas)

        df_metricas["Seccion"] = df_metricas["Seccion"].astype(
            str).str.zfill(11)
        gdf_join = gdf_secciones.merge(
            df_metricas, left_on="MUNDISSEC", right_on="Seccion", how="inner"
        )

        # 3. Definir métricas a mapear

        metricas_mapas = [
            ("Precision (%)", "Precision (%)"),
            ("Recall (%)", "Recall (%)"),
            ("F1_Score", "F1-Score"),
        ]

        fig, axes = plt.subplots(1, 3, figsize=(20, 6))
        fig.suptitle(
            "Mapas temáticos post-validación manual", fontsize=16, fontweight="bold"
        )

        for i, (col, titulo) in enumerate(metricas_mapas):
            ax = axes[i]
            gdf_join.plot(
                column=col,
                ax=ax,
                legend=True,
                cmap="RdYlGn",
                edgecolor="black",
                linewidth=0.5,
                missing_kwds={"color": "lightgrey", "label": "No validado"},
            )
            ax.set_title(titulo, fontsize=14)
            ax.set_axis_off()
        plt.tight_layout(rect=[0, 0.03, 1, 0.95])
        salida = self.config.output_dir / "mapas_tematicos_metricas_post_validacion.png"
        plt.savefig(salida, dpi=300, bbox_inches="tight")
        plt.close()
        print(f"✅ Mapas temáticos post-validación guardados en {salida}")

    def generar_mapa_fiabilidad_final(self, df_consolidado: pd.DataFrame) -> None:
        """
        Genera un mapa temático con TODAS las secciones censales (inalterables),
        incluyendo aquellas no analizadas, con una categoría adicional: 'No analizada'.
        También exporta un shapefile por cada categoría de fiabilidad.
        """
        import matplotlib.pyplot as plt
        from pathlib import Path

        # Preparar datos
        gdf = self.shp_secciones.copy()
        gdf["MUNDISSEC"] = gdf["MUNDISSEC"].astype(str).str.zfill(11)
        df_consolidado["Sección"] = df_consolidado["Sección"].astype(str).str.zfill(11)

        # Merge y limpieza de datos
        gdf = gdf.merge(df_consolidado[["Sección", "Decisión"]], left_on="MUNDISSEC", right_on="Sección", how="left")
        gdf["Decisión"] = gdf["Decisión"].fillna("No analizada").str.replace(r"[✅❌⚠️]", "", regex=True).str.strip()

        # Estadísticas
        total_secciones = len(gdf)
        conteos = gdf["Decisión"].value_counts()
        porcentajes = (conteos / total_secciones * 100).round(1)

        # Exportar shapefiles por categoría
        carpeta_salida = self.config.output_dir / "shapefiles_decision_fiabilidad"
        carpeta_salida.mkdir(parents=True, exist_ok=True)
        
        for decision in gdf["Decisión"].dropna().unique():
            subset = gdf[gdf["Decisión"] == decision]
            ruta_salida = carpeta_salida / f"secciones_{decision.lower().replace(' ', '_')}_{len(subset)}.shp"
            subset.to_file(ruta_salida, driver="ESRI Shapefile")
            print(f"📁 Exportado: {ruta_salida}")

        # Generar mapa
        colores = {"Fiable": "#5cb85c", "Condicional": "#f0ad4e", "No fiable": "#d9534f", "No analizada": "#dcdcdc"}
        
        fig, ax = plt.subplots(1, 1, figsize=(16, 10))
        handles, labels = [], []

        for decision, color in colores.items():
            subset = gdf[gdf["Decisión"] == decision]
            if not subset.empty:
                subset.plot(ax=ax, color=color, edgecolor="black", linewidth=0.3)
                count, pct = conteos.get(decision, 0), porcentajes.get(decision, 0.0)
                handles.append(plt.Rectangle((0, 0), 1, 1, facecolor=color, edgecolor='black', linewidth=0.5))
                labels.append(f"{decision}\n({count:,} secciones - {pct}%)")

        # Configurar mapa
        ax.set_title("Mapa de fiabilidad por secciones censales", fontsize=16, fontweight='bold', pad=20)
        ax.set_axis_off()

        # Leyenda
        legend = ax.legend(handles, labels, title="Clasificación", loc="center left", bbox_to_anchor=(1.02, 0.5),
                        fontsize=11, title_fontsize=12, frameon=True, fancybox=True, shadow=False, framealpha=0.9)
        legend.get_title().set_fontweight('bold')

        # Texto resumen
        total_analizadas = total_secciones - conteos.get("No analizada", 0)
        ax.text(0.02, 0.02, f"Total de secciones: {total_secciones:,}\nSecciones analizadas: {total_analizadas:,}",
                transform=ax.transAxes, fontsize=10, verticalalignment='bottom',
                bbox=dict(boxstyle="round,pad=0.5", facecolor="white", alpha=0.8))

        # Guardar y mostrar estadísticas
        output_path = self.config.output_dir / "mapa_decision_final_completo.png"
        plt.savefig(output_path, dpi=300, bbox_inches="tight", facecolor='white', edgecolor='none')
        plt.close()

        print(f"✅ Mapa completo de decisión guardado en: {output_path}")
        print("\n📊 Estadísticas por categoría:")
        for decision in ["Fiable", "Condicional", "No fiable", "No analizada"]:
            if decision in conteos:
                print(f"   {decision}: {conteos[decision]:,} secciones ({porcentajes[decision]}%)")
        print(f"\n📍 Total de secciones: {total_secciones:,}")
        print(f"📍 Secciones analizadas: {total_analizadas:,}")

    def resumen_registros_a_validar(self):
        """Genera resumen de registros a validar con autoajuste de columnas en Excel"""

        # Buscar todos los archivos de validación
        archivos = list(self.config.output_dir.glob("**/validacion_visual_unificada_*.xlsx")) + \
            list(self.config.output_dir.glob(
                "**/dummy_validacion_visual_unificada_*.xlsx"))

        total_registros = 0
        resumen = []
        conteo_mundissec = {}

        # Procesar archivos
        for archivo in archivos:
            try:
                df = pd.read_excel(archivo)
                n_registros = len(df)
                total_registros += n_registros

                # Obtener códigos MUNDISSEC
                if "MUNDISSEC" in df.columns:
                    codigos_serie = df["MUNDISSEC"].astype(str).str.zfill(11)
                    codigos = codigos_serie.unique().tolist()
                    # Actualizar conteo
                    for cod, count in codigos_serie.value_counts().items():
                        conteo_mundissec[cod] = conteo_mundissec.get(
                            cod, 0) + count
                else:
                    # Extraer del nombre del archivo
                    match = re.search(r"seccion_(\d{11})", archivo.name)
                    codigos = [match.group(1)] if match else []
                    if codigos:
                        conteo_mundissec[codigos[0]] = conteo_mundissec.get(
                            codigos[0], 0) + n_registros

                resumen.append({
                    "Archivo": archivo.name,
                    "Registros a validar": n_registros,
                    "MUNDISSEC(s)": ",".join(codigos),
                })

            except Exception as e:
                logger.warning(f"Error leyendo {archivo}: {e}")

        logger.info(
            f"🔎 Total registros a validar visualmente: {total_registros}")

        # Crear DataFrame y Excel
        df_resumen = pd.DataFrame(resumen)
        df_resumen["¿Validado?"] = ""
        excel_path = self.config.output_dir / "resumen_registros_a_validar.xlsx"

        # Escribir Excel y configurar validación
        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            df_resumen.to_excel(writer, sheet_name="Resumen", index=False)

        # Configurar validación y formato
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["Resumen"]

        # Encontrar columna '¿Validado?' y configurar validación
        col_validado = next((i for i, cell in enumerate(
            ws[1], 1) if cell.value == "¿Validado?"), None)
        if not col_validado:
            raise Exception(
                "No se encontró la columna '¿Validado?' en el Excel generado.")

        col_letter = get_column_letter(col_validado)
        rango = f"{col_letter}2:{col_letter}{ws.max_row}"

        # Añadir validación de lista
        dv = DataValidation(type="list", formula1='"✔️,❌"', allow_blank=True)
        dv.error = "Selecciona ✔️ o ❌"
        dv.prompt = "Marca si está validado"
        ws.add_data_validation(dv)
        dv.add(rango)

        # Autoajustar columnas
        for i, col in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), 1):
            max_length = max(len(str(cell.value)) for cell in col)
            ws.column_dimensions[get_column_letter(
                i)].width = min(max_length + 2, 50)

        wb.save(excel_path)

        # Generar resumen en texto
        txt_path = self.config.output_dir / "resumen_registros_a_validar.txt"
        with open(txt_path, "w", encoding="utf-8") as f:
            f.write(f"Total registros a validar: {total_registros}\n\n")
            f.write("Secciones involucradas (MUNDISSEC):\n")

            for cod in sorted(conteo_mundissec.keys()):
                f.write(f"- {cod}: {conteo_mundissec[cod]} registros\n")

            f.write("\nDetalle por archivo:\n")
            for reg in resumen:
                f.write(
                    f"{reg['Archivo']}: {reg['Registros a validar']} ({reg['MUNDISSEC(s)']})\n")



def main():
    """Función principal para ejecutar el análisis"""
    # Configuración

    config = Config()

    # Crear analizador

    analyzer = OptimizedCyclewayAnalyzer(config)

    # Detectar fase a ejecutar

    if len(sys.argv) > 1:
        fase = sys.argv[1].lower()
        if fase not in ["auto", "diagnostico", "post_validacion", "forzar_diagnostico"]:
            print(
                "Uso: python script.py [auto|diagnostico|post_validacion|forzar_diagnostico]"
            )
            print("  auto: Detecta automáticamente la fase (por defecto)")
            print("  diagnostico: Ejecuta Fase 1")
            print("  post_validacion: Ejecuta Fase 2")
            print("  forzar_diagnostico: Re-ejecuta Fase 1 (sobreescribe archivos)")
            return
    else:
        fase = "auto"  # Por defecto usar detección automática
    try:
        # Ejecutar análisis según la fase

        analyzer.ejecutar_analisis_completo(fase=fase)

        if fase in ["diagnostico", "forzar_diagnostico"]:
            print("\n" + "=" * 70)
            print("🎯 FASE 1 COMPLETADA EXITOSAMENTE")
            print("=" * 70)
            print("📋 PRÓXIMOS PASOS:")
            print("1. Revisar archivos 'validacion_visual_unificada_*.xlsx'")
            print("2. Completar columna '¿Es realmente un carril bici?'")
            print("   - ✅ Confirmado: Es un carril bici real")
            print("   - ❌ Falso Positivo: No es un carril bici")
            print("   - ❌ Falso Negativo: Debería estar en OSM")
            print("3. Ejecutar: python script.py post_validacion")
            print("   O simplemente: python script.py (detección automática)")
            print("=" * 70)
        elif fase == "post_validacion":
            print("\n" + "=" * 70)
            print("🏁 ANÁLISIS COMPLETO FINALIZADO")
            print("=" * 70)
            print("📊 Revise los reportes finales en:")
            print("- metricas_validadas_post_GSV.xlsx")
            print("- media_resumen_metricas_validados.xlsx")
            print("- graficos_resultados_finales.png")
            print("=" * 70)
    except KeyboardInterrupt:
        logger.info("Análisis interrumpido por el usuario")
    except Exception as e:
        logger.error(f"Error durante el análisis: {e}")
        raise


if __name__ == "__main__":
    main()
